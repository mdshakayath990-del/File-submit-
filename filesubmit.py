import telebot
from telebot import types
import re
import logging
import os
import json
from datetime import datetime
import openpyxl
from io import BytesIO
import time
import warnings

# Warning মেসেজ বন্ধ করার ম্যাজিক ট্রিক
warnings.filterwarnings("ignore")

# ════════════════════════════════════════════════════════════════
#  CONFIGURATION  —  শুধু এই অংশ পরিবর্তন করুন
# ════════════════════════════════════════════════════════════════
BOT_TOKEN      = os.environ.get("BOT_TOKEN", "8540616234:AAGEtDCYh77B0VQIlEFek57my45S_GtnrS8")

# ⚠️ গুরুত্বপূর্ণ: বটকে চ্যানেলে মেসেজ পাঠাতে হলে Numerical ID (-100...) লাগবে। 
CHANNEL_ID     = -1003945398035 

ADMIN_IDS      = [7689218221, 7833093821]
OWNER_SUPPORT  = "@SKYSMSOWNER"      # অনার সাপোর্ট ইউজারনেম দিন
ADMIN_SUPPORT  = "@FBSKYSUPPORT"     # এডমিন সাপোর্ট ইউজারনেম দিন
BOT_NAME       = "𝐅𝐁 𝐈𝐃 𝐒𝐮𝐛𝐦𝐢𝐭 𝐇𝐮𝐛"
MAX_HISTORY    = 10

# STYLIZED BUTTON LABELS & STRINGS
BTN_SELL      = "♂️ 𝐒𝐄𝐋𝐋 🆔"
BTN_PRICE     = "📊 𝐏𝐫𝐢𝐜𝐞 𝐋𝐢𝐬𝐭"
BTN_SUPPORT   = "🎧 𝐒𝐮𝐩𝐩𝐨𝐫𝐭"
BTN_HISTORY   = "📜 𝐌𝐲 𝐇𝐢𝐬𝐭𝐨𝐫𝐲"
BTN_PROFILE   = "👤 𝐌𝐲 𝐏𝐫𝐨𝐟𝐢𝐥𝐞"
BTN_NOTICE    = "📣 𝐍𝐨𝐭𝐢𝐜𝐞 𝐁𝐨𝐚𝐫𝐝"
BTN_ADMIN     = "🔐 𝐀𝐝𝐦𝐢𝐧 𝐏𝐚𝐧𝐞𝐥"
BTN_BROADCAST = "📢 𝐁𝐫𝐨𝐚𝐝𝐜𝐚𝐬𝐭"

MENU_BUTTONS = [BTN_SELL, BTN_PRICE, BTN_SUPPORT, BTN_HISTORY, BTN_PROFILE, BTN_NOTICE, BTN_ADMIN, BTN_BROADCAST]

SYSTEM_SETTINGS = {
    "min_id_limit": 5,           
    "maintenance_mode": False,
    "welcome_msg": f"📌 *কীভাবে 𝐈𝐃 জমা দেবেন:*\n1️⃣ *{BTN_SELL}* বাটন চাপুন\n2️⃣ 𝐂𝐚𝐭𝐞𝐠𝐨𝐫𝐲 সিলেক্ট করুন\n3️⃣ 𝐄𝐱𝐜𝐞𝐥 (.𝐱𝐥𝐬𝐱) ফাইল আপলোড করুন\n4️⃣ আপনার 𝐓𝐞𝐥𝐞𝐠𝐫𝐚𝐦 𝐮𝐬𝐞𝐫𝐧𝐚𝐦𝐞 দিন\n5️⃣ 𝐏𝐚𝐲𝐦𝐞𝐧𝐭 মেথড ও নম্বর দিন\n✅ সাবমিট সম্পন্ন — 𝐀𝐝𝐦𝐢𝐧 𝐫𝐞𝐬𝐮𝐥𝐭 জানাবে",
    "notice_board": "",
    "pay_bkash": True,  
    "pay_nagad": True   
}

# ════════════════════════════════════════════════════════════════
#  CATEGORIES
# ════════════════════════════════════════════════════════════════
CATEGORIES = {
    "PC1000X":  {"name": "📲 𝐏𝐂 𝐂𝐥𝐨𝐧𝐞 𝟏𝟎𝟎𝟎𝐱",         "rate": 13.00, "format": "𝐀=𝐔𝐈𝐃 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝐂𝐨𝐨𝐤𝐢𝐞",              "status": "open"},
    "PC6155X":  {"name": "📲 𝐏𝐂 𝐂𝐥𝐨𝐧𝐞 𝟔𝟏𝟓𝟓𝐱/𝟓𝟔𝐱/𝟓𝟕𝐱", "rate":  7.00, "format": "𝐀=𝐔𝐈𝐃 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝐂𝐨𝐨𝐤𝐢𝐞",              "status": "open"},
    "PC6158X":  {"name": "📲 𝐏𝐂 𝐂𝐥𝐨𝐧𝐞 𝟔𝟏𝟓𝟖𝐱",         "rate":  4.00, "format": "𝐀=𝐔𝐈𝐃 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝐂𝐨𝐨𝐤𝐢𝐞",              "status": "open"},
    "NUM2FA":   {"name": "⚡ 𝐍𝐮𝐦𝐛𝐞𝐫 𝟐𝐅𝐀 𝐈'𝐃",          "rate":  6.00, "format": "𝐀=𝐔𝐈𝐃 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝟐𝐅𝐀 𝐊𝐞𝐲",             "status": "open"},
    "NUMCOOKIE":{"name": "📞 𝐍𝐮𝐦𝐛𝐞𝐫 𝐂𝐨𝐨𝐤𝐢𝐞𝐬 𝐈'𝐃",      "rate":  4.00, "format": "𝐀=𝐔𝐈𝐃 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝐂𝐨𝐨𝐤𝐢𝐞",              "status": "open"},
    "HOTMAIL30":{"name": "☁️ 𝐇𝐨𝐭𝐦𝐚𝐢𝐥 𝟑𝟎+ 𝐅𝐫𝐢𝐞𝐧𝐝",      "rate": 10.00, "format": "𝐀=𝐔𝐈𝐃 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝟐𝐅𝐀 | 𝐃=𝐅𝐮𝐥𝐥 𝐌𝐚𝐢𝐥",   "status": "open"},
    "HOTMAIL00":{"name": "💠 𝐇𝐨𝐭𝐦𝐚𝐢𝐥 𝟎𝟎 𝐅𝐫𝐢𝐞𝐧𝐝",       "rate":  7.00, "format": "𝐀=𝐔𝐈𝐃 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝟐𝐅𝐀 | 𝐃=𝐅𝐮𝐥𝐥 𝐌𝐚𝐢𝐥",   "status": "open"},
    "INSTA2FA": {"name": "📸 𝐈𝐧𝐬𝐭𝐚𝐠𝐫𝐚𝐦 𝟐𝐅𝐀",           "rate":  2.70, "format": "𝐀=𝐔𝐬𝐞𝐫𝐧𝐚𝐦𝐞 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 | 𝐂=𝟐𝐅𝐀 𝐊𝐞𝐲",        "status": "open"},
    "INSTACOOK":{"name": "🍪 𝐈𝐧𝐬𝐭𝐚𝐠𝐫𝐚𝐦 𝐂𝐨𝐨𝐤𝐢𝐞𝐬",       "rate":  4.00, "format": "𝐀=𝐔𝐬𝐞𝐫𝐧𝐚𝐦𝐞 | 𝐁=𝐏𝐚𝐬𝐬𝐰𝐨𝐫𝐝 (𝟐 𝐜𝐨𝐥𝐮𝐦𝐧𝐬 𝐨𝐧𝐥𝐲)",  "status": "open"},
}

# ════════════════════════════════════════════════════════════════
#  INIT
# ════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger(__name__)

bot = telebot.TeleBot(BOT_TOKEN, parse_mode=None)

user_data        = {}
all_submissions  = {}
user_submissions = {}
submission_count = 0
registered_users = set()
BANNED_USERS     = set()
username_to_id   = {}

# ════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════
def is_admin(cid):        
    return cid in ADMIN_IDS

def is_banned(cid):
    return cid in BANNED_USERS

def safe_md(text):
    escape_chars = r'_*[]()~`>#+-=|{}.!'
    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', str(text))

def main_menu(cid):
    m = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    m.add(types.KeyboardButton(BTN_SELL))
    m.row(types.KeyboardButton(BTN_PRICE), types.KeyboardButton(BTN_HISTORY))
    m.row(types.KeyboardButton(BTN_SUPPORT), types.KeyboardButton(BTN_PROFILE))
    m.row(types.KeyboardButton(BTN_NOTICE))
    
    if is_admin(cid):
        m.row(types.KeyboardButton(BTN_ADMIN), types.KeyboardButton(BTN_BROADCAST))
    return m

def cancel_btn():
    m = types.InlineKeyboardMarkup()
    m.add(types.InlineKeyboardButton("❌ 𝐂𝐀𝐍𝐂𝐄𝐋", callback_data="cancel_flow"))
    return m

def build_price_list():
    lines = [
        f"🏷 *{BTN_PRICE}*",
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    ]
    for cat in CATEGORIES.values():
        st = "✅ 𝐎𝐩𝐞𝐧" if cat["status"] == "open" else "🔴 𝐂𝐥𝐨𝐬𝐞𝐝"
        lines.append(
            f"\n{'✅' if cat['status']=='open' else '🔴'} *{cat['name']}*\n"
            f"   💰 𝐑𝐚𝐭𝐞: *{cat['rate']:.2f} 𝐓𝐤/𝐩𝐜𝐬*\n"
            f"   📋 `{cat['format']}`\n"
            f"   𝐒𝐭𝐚𝐭𝐮𝐬: {st}"
        )
    lines.append("\n━━━━━━━━━━━━━━━━━━━━━━━━")
    return "\n".join(lines)

def register_user(message):
    registered_users.add(message.chat.id)
    if message.from_user and message.from_user.username:
        username_to_id[f"@{message.from_user.username.lower()}"] = message.chat.id

def get_admin_channel_markup(sub_id, is_completed=False, live=0):
    sub = all_submissions.get(sub_id, {})
    rcv_mark = "✅" if sub.get("rcv_status") else "❌"
    pay_mark = "✅" if sub.get("pay_status") else "❌"
    
    m = types.InlineKeyboardMarkup(row_width=2)
    m.row(
        types.InlineKeyboardButton(f"📥 𝐑𝐞𝐜𝐞𝐢𝐯𝐞𝐝: {rcv_mark}", callback_data=f"admrcv_{sub_id}"),
        types.InlineKeyboardButton(f"💸 𝐏𝐚𝐲𝐦𝐞𝐧𝐭: {pay_mark}", callback_data=f"admpay_{sub_id}")
    )
    if is_completed:
        m.add(types.InlineKeyboardButton(f"✅ 𝐂𝐨𝐦𝐩𝐥𝐞𝐭𝐞𝐝 — {live} 𝐏𝐚𝐢𝐝", callback_data="already_done"))
    else:
        m.add(types.InlineKeyboardButton("🟢 𝐑𝐞𝐯𝐢𝐞𝐰 / 𝐑𝐞𝐬𝐮𝐥𝐭 দিন", callback_data=f"review_{sub_id}"))
    return m

# ════════════════════════════════════════════════════════════════
#  /start
# ════════════════════════════════════════════════════════════════
@bot.message_handler(commands=["start"])
def cmd_start(message):
    cid = message.chat.id
    if is_banned(cid): return

    if SYSTEM_SETTINGS["maintenance_mode"] and not is_admin(cid):
        bot.send_message(cid, "🛠️ সিস্টেম বর্তমানে আপডেটের কাজে সাময়িকভাবে বন্ধ আছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।")
        return

    register_user(message)
    name = message.from_user.first_name or "বন্ধু"

    welcome = (
        f"👋 আসসালামু আলাইকুম, *{name}*!\n\n"
        f"🔥 *{BOT_NAME}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"{SYSTEM_SETTINGS['welcome_msg']}\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    bot.send_message(cid, welcome, reply_markup=main_menu(cid), parse_mode="Markdown")

# ════════════════════════════════════════════════════════════════
#  MAIN TEXT HANDLER
# ════════════════════════════════════════════════════════════════
@bot.message_handler(func=lambda m: True, content_types=["text"])
def handle_text(message):
    cid  = message.chat.id
    
    if is_banned(cid): return
    
    if SYSTEM_SETTINGS["maintenance_mode"] and not is_admin(cid):
        bot.send_message(cid, "🛠️ সিস্টেম বর্তমানে আপডেটের কাজে সাময়িকভাবে বন্ধ আছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।")
        return

    register_user(message)
    text = (message.text or "").strip()

    if text == BTN_SELL:
        show_categories(cid)

    elif text == BTN_PRICE:
        bot.send_message(cid, build_price_list(), parse_mode="Markdown")

    elif text == BTN_SUPPORT:
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn_owner = types.InlineKeyboardButton("👑 𝐎𝐰𝐧𝐞𝐫 𝐒𝐮𝐩𝐩𝐨𝐫𝐭", url=f"https://t.me/{OWNER_SUPPORT.replace('@', '')}")
        btn_admin = types.InlineKeyboardButton("👨‍💻 𝐀𝐝𝐦𝐢𝐧 𝐒𝐮𝐩𝐩𝐨𝐫𝘁", url=f"https://t.me/{ADMIN_SUPPORT.replace('@', '')}")
        markup.add(btn_owner, btn_admin)
        bot.send_message(cid, "🎧 *𝐒𝐮𝐩𝐩𝐨𝐫𝐭 𝐂𝐞𝐧𝐭𝐞𝐫*\n━━━━━━━━━━━━━━━━━━━━━━━━\nযেকোনো প্রয়োজনে নিচের বাটনে ক্লিক করে সরাসরি যোগাযোগ করুন:", reply_markup=markup, parse_mode="Markdown")

    elif text == BTN_HISTORY:
        show_history(cid)

    elif text == BTN_PROFILE:
        show_profile(cid)

    elif text == BTN_NOTICE:
        notice = SYSTEM_SETTINGS.get("notice_board", "")
        if not notice:
            bot.send_message(cid, f"📣 *{BTN_NOTICE}*\n━━━━━━━━━━━━━━━━━━━━━━━━\n_এখনো কোনো নোটিশ নেই।_", parse_mode="Markdown")
        else:
            bot.send_message(cid, f"📣 *{BTN_NOTICE}*\n━━━━━━━━━━━━━━━━━━━━━━━━\n{notice}", parse_mode="Markdown")

    elif text == BTN_ADMIN:
        if is_admin(cid):
            show_admin_panel(cid)
            
    elif text == BTN_BROADCAST:
        if is_admin(cid):
            msg = bot.send_message(cid, "📢 *𝐁𝐫𝐨𝐚𝐝𝐜𝐚𝐬𝐭 মেসেজ লিখুন*\n(আপনি টেক্সট, ছবি বা স্টিকারও পাঠাতে পারেন):", parse_mode="Markdown", reply_markup=cancel_btn())
            bot.register_next_step_handler(msg, step_broadcast)

    else:
        bot.send_message(cid, "❓ বুঝতে পারিনি। নিচের বাটনগুলো ব্যবহার করুন।", reply_markup=main_menu(cid))

def show_profile(cid):
    subs = user_submissions.get(cid, [])
    total_submitted = len(subs)
    total_approved = 0
    total_earned = 0.0

    for sid in subs:
        s = all_submissions.get(sid, {})
        if s.get("paid"):
            total_approved += s.get("live_qty", 0)
            total_earned += s.get("total", 0.0)

    profile_text = (
        f"👤 *{BTN_PROFILE}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 𝐔𝐬𝐞𝐫 𝐈𝐃: `{cid}`\n"
        f"📦 মোট সাবমিট: *{total_submitted}* বার\n"
        f"✅ মোট এপ্রুভড আইডি: *{total_approved}* টি\n"
        f"💰 মোট আয়: *৳{total_earned:.2f}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    bot.send_message(cid, profile_text, parse_mode="Markdown")

# ════════════════════════════════════════════════════════════════
#  SELL FLOW
# ════════════════════════════════════════════════════════════════
def show_categories(cid):
    markup = types.InlineKeyboardMarkup(row_width=1)
    for code, cat in CATEGORIES.items():
        if cat["status"] == "open":
            label = f"✅ {cat['name']}  —  {cat['rate']:.2f} 𝐓𝐤"
        else:
            label = f"🔴 {cat['name']}  [বন্ধ]"
        markup.add(types.InlineKeyboardButton(label, callback_data=f"sell_{code}"))

    bot.send_message(cid, "🛒 *𝐈𝐃 বিক্রি করুন*\n━━━━━━━━━━━━━━━━━━━━━━━━\n\n👇 *𝐂𝐚𝐭𝐞𝐠𝐨𝐫𝐲 সিলেক্ট করুন:*", reply_markup=markup, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("sell_"))
def cb_sell(call):
    bot.answer_callback_query(call.id)
    cid  = call.message.chat.id
    code = call.data[5:]

    if code not in CATEGORIES: return
    cat = CATEGORIES[code]
    if cat["status"] != "open":
        bot.send_message(cid, "🔴 এই ক্যাটাগরি বর্তমানে বন্ধ আছে।")
        return

    user_data[cid] = {
        "type": code, "type_name": cat["name"],
        "rate": cat["rate"], "format": cat["format"],
        "date": datetime.now().strftime("%d %b %Y"),
        "step": "file" 
    }

    msg = bot.send_message(
        cid,
        f"✅ *𝐒𝐞𝐥𝐞𝐜𝐭𝐞𝐝:* {cat['name']}\n💰 𝐑𝐚𝐭𝐞: *{cat['rate']:.2f} 𝐓𝐤/𝐩𝐜𝐬*\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📤 *এখন 𝐄𝐱𝐜𝐞𝐥 ফাইল পাঠান*\n📋 𝐂𝐨𝐥𝐮𝐦𝐧 𝐋𝐚𝐲𝐨𝐮𝐭:\n`{cat['format']}`\n\n"
        f"⚠️ শুধুমাত্র *.𝐱𝐥𝐬𝐱* ফাইল সাপোর্ট করে",
        reply_markup=cancel_btn(), parse_mode="Markdown"
    )
    bot.register_next_step_handler(msg, step_file_text)

def step_file_text(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if message.content_type == "document": return handle_docs(message)
    bot.send_message(cid, "❌ একটি *.𝐱𝐥𝐬𝐱* ফাইল পাঠান।", reply_markup=cancel_btn(), parse_mode="Markdown")
    bot.register_next_step_handler(message, step_file_text)

@bot.message_handler(content_types=["document"])
def handle_docs(message):
    cid = message.chat.id
    if is_banned(cid): return
    if cid not in user_data or user_data[cid].get("step") != "file": return

    fname = message.document.file_name or ""
    if not fname.lower().endswith(".xlsx"):
        msg = bot.reply_to(message, "❌ শুধু *.𝐱𝐥𝐬𝐱* ফাইল সাপোর্ট করে!\nএকটি 𝐄𝐱𝐜𝐞𝐥 ফাইল পাঠান:", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_file_text)

    try:
        fi   = bot.get_file(message.document.file_id)
        raw  = bot.download_file(fi.file_path)
        wb   = openpyxl.load_workbook(filename=BytesIO(raw), data_only=True)
        ws   = wb.active
        
        uids = []
        valid_rows = 0
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                uid = str(row[0]).strip() if row and row[0] is not None else ""
                if uid: 
                    uids.append(uid)
                valid_rows += 1

        qty = valid_rows
        min_limit = SYSTEM_SETTINGS["min_id_limit"]
        
        if qty < min_limit:
            msg = bot.reply_to(message, f"❌ আপনি মাত্র *{qty}* টি 𝐈𝐃 দিয়েছেন।\n⚠️ মিনিমাম *{min_limit}* টি 𝐈𝐃 একসাথে দিতে হবে!\n\nসঠিক ফাইলটি আবার পাঠান:", reply_markup=cancel_btn(), parse_mode="Markdown")
            return bot.register_next_step_handler(msg, step_file_text)

        # -------------------------------------------------------------
        #  VALIDATION
        # -------------------------------------------------------------
        if len(uids) != len(set(uids)):
            msg = bot.reply_to(message, "❌ *ফাইলে ডুপ্লিকেট (একই আইডি একাধিকবার) পাওয়া গেছে!*\nঅনুগ্রহ করে ডুপ্লিকেট রিমুভ করে সম্পূর্ণ ফ্রেশ ফাইলটি আবার দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
            return bot.register_next_step_handler(msg, step_file_text)
            
        category_code = user_data[cid]["type"]
        if category_code == "PC1000X" and any(not str(u).startswith("1000") for u in uids):
            msg = bot.reply_to(message, "❌ *ভুল ক্যাটাগরি!*\nআপনি 𝟏𝟎𝟎𝟎𝐱 সিলেক্ট করেছেন কিন্তু ফাইলে অন্য 𝐈𝐃 (যেমন 𝟔𝟏𝐱) পাওয়া গেছে। সঠিক ফাইল দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
            return bot.register_next_step_handler(msg, step_file_text)
            
        elif category_code == "PC6155X" and any(not (str(u).startswith("6155") or str(u).startswith("6156") or str(u).startswith("6157")) for u in uids):
             msg = bot.reply_to(message, "❌ *ভুল ক্যাটাগরি!*\nআপনি 𝟔𝟏𝟓𝟓𝐱/𝟓𝟔𝐱/𝟓𝟕𝐱 সিলেক্ট করেছেন কিন্তু ফাইলে অন্য 𝐈𝐃 পাওয়া গেছে। সঠিক ফাইল দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
             return bot.register_next_step_handler(msg, step_file_text)
             
        elif category_code == "PC6158X" and any(not str(u).startswith("6158") for u in uids):
             msg = bot.reply_to(message, "❌ *ভুল ক্যাটাগরি!*\nআপনি 𝟔𝟏𝟓𝟖𝐱 সিলেক্ট করেছেন কিন্তু ফাইলে অন্য 𝐈𝐃 পাওয়া গেছে। সঠিক ফাইল দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
             return bot.register_next_step_handler(msg, step_file_text)

        user_data[cid].update({"qty": qty, "file_name": fname, "file_id": message.document.file_id, "step": "username"})

        msg = bot.reply_to(
            message,
            f"✅ *𝐅𝐢𝐥𝐞 𝐀𝐜𝐜𝐞𝐩𝐭𝐞𝐝 (𝐍𝐨 𝐃𝐮𝐩𝐥𝐢𝐜𝐚𝐭𝐞𝐬)*\n📊 𝐓𝐨𝐭𝐚𝐥 𝐑𝐨𝐰𝐬: *{qty} 𝐩𝐜𝐬*\n\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "👤 *আপনার 𝐓𝐞𝐥𝐞𝐠𝐫𝐚𝐦 𝐮𝐬𝐞𝐫𝐧𝐚𝐦𝐞 দিন:*\n_উদাহরণ: @myusername_",
            reply_markup=cancel_btn(), parse_mode="Markdown"
        )
        bot.register_next_step_handler(msg, step_username)

    except Exception as e:
        log.error(f"File error: {e}")
        msg = bot.reply_to(message, f"❌ ফাইল পড়তে সমস্যা: `{e}`\nআবার চেষ্টা করুন:", reply_markup=cancel_btn(), parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_file_text)


def step_username(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    uname = (message.text or "").strip()
    if not re.match(r'^@[a-zA-Z0-9_]{4,32}$', uname):
        msg = bot.send_message(cid, "❌ *ভুল ফরম্যাট!*\n@ দিয়ে শুরু করুন\n_উদাহরণ: @myusername_", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_username)

    username_to_id[uname.lower()] = cid
    user_data[cid]["user"] = uname
    user_data[cid]["step"] = "pay_method"

    m = types.InlineKeyboardMarkup(row_width=2)
    btns = []
    if SYSTEM_SETTINGS.get("pay_bkash", True):
        btns.append(types.InlineKeyboardButton("🟣 𝐁𝐤𝐚𝐬𝐡", callback_data="pay_bkash"))
    if SYSTEM_SETTINGS.get("pay_nagad", True):
        btns.append(types.InlineKeyboardButton("🟠 𝐍𝐚𝐠𝐚𝐝", callback_data="pay_nagad"))

    if not btns:
        bot.send_message(cid, "❌ বর্তমানে পেমেন্ট গ্রহণ বন্ধ আছে। 𝐀𝐝𝐦𝐢𝐧 কে জানান।", reply_markup=main_menu(cid))
        del user_data[cid]
        return

    m.add(*btns)
    m.add(types.InlineKeyboardButton("❌ 𝐂𝐀𝐍𝐂𝐄𝐋", callback_data="cancel_flow"))

    bot.send_message(
        cid,
        "💳 *𝐏𝐚𝐲𝐦𝐞𝐧𝐭 𝐌𝐞𝐭𝐡𝐨𝐝 সিলেক্ট করুন:*",
        reply_markup=m, parse_mode="Markdown"
    )

def step_number(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    number = (message.text or "").strip()
    if not re.match(r'^01[0-9]{9}$', number):
        msg = bot.send_message(cid, "❌ *ভুল নম্বর!* ১১ ডিজিটের বাংলাদেশী নম্বর দিন\n_উদাহরণ: 01712345678_", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_number)

    method_name = user_data[cid].get("pay_method_name", "Unknown")
    user_data[cid]["account"] = f"{method_name} - {number}"
    
    # নোট অপশন বাদ দিয়ে সরাসরি ফাইনাল সাবমিশন করা হচ্ছে
    finalize(cid, "𝐍/𝐀")

def finalize(cid, note):
    global submission_count
    submission_count += 1
    sub_id = f"SUB{submission_count:04d}"
    d      = user_data[cid]
    est    = d["qty"] * d["rate"]

    all_submissions[sub_id] = {
        "chat_id":   cid,
        "user":      d["user"],
        "type_name": d["type_name"],
        "rate":      d["rate"],
        "qty":       d["qty"],
        "total":     0,
        "account":   d["account"],
        "note":      note,
        "date":      d["date"],
        "paid":      False,
        "live_qty":  0,
        "file_name": d["file_name"],
        "rcv_status": False,
        "pay_status": False
    }
    user_submissions.setdefault(cid, []).append(sub_id)

    receipt = (
        "✅ *𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 𝐒𝐮𝐜𝐜𝐞𝐬𝐬𝐟𝐮𝐥!*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 𝐒𝐮𝐛 𝐈𝐃  : `{sub_id}`\n📁 𝐓𝐲𝐩𝐞    : {d['type_name']}\n📅 𝐃𝐚𝐭𝐞    : {d['date']}\n📄 𝐅𝐢𝐥𝐞    : `{d['file_name']}`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 𝐓𝐨𝐭𝐚𝐥   : *{d['qty']} 𝐩𝐜𝐬*\n💰 𝐄𝐬𝐭. 𝐀𝐦𝐭 : *{est:.2f} 𝐓𝐤* ({d['rate']} × {d['qty']})\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💳 𝐀𝐜𝐜𝐨𝐮𝐧𝐭 : `{d['account']}`\n📝 𝐍𝐨𝐭𝐞    : {note}\n\n"
        "⏳ *𝐀𝐝𝐦𝐢𝐧 𝐫𝐞𝐯𝐢𝐞𝐰 করার পর 𝐫𝐞𝐬𝐮𝐥𝐭 পাঠাবে।*"
    )

    adm_cap = (
        f"📥 *𝐍𝐞𝐰 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 #{sub_id}*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 𝐔𝐬𝐞𝐫    : {safe_md(d['user'])} | `{cid}`\n"
        f"🏷️ 𝐓𝐲𝐩𝐞    : {safe_md(d['type_name'])}\n"
        f"📊 𝐑𝐨𝐰𝐬    : *{d['qty']} 𝐩𝐜𝐬*\n"
        f"💰 𝐄𝐬𝐭.    : *{est:.2f} 𝐓𝐤*\n"
        f"💳 𝐀𝐜𝐜𝐨𝐮𝐧𝐭 : `{safe_md(d['account'])}`\n"
        f"📝 𝐍𝐨𝐭𝐞    : {safe_md(note)}"
    )
    
    adm_btn = get_admin_channel_markup(sub_id)

    bot.send_message(cid, receipt, parse_mode="Markdown", reply_markup=main_menu(cid))
    try:
        bot.send_document(CHANNEL_ID, d["file_id"], caption=adm_cap, parse_mode="Markdown", reply_markup=adm_btn)
    except Exception as e: 
        log.error(f"Channel send error: {e}")
        bot.send_message(cid, "⚠️ ফাইল জমা হয়েছে কিন্তু চ্যানেলে ফরোয়ার্ড হতে সমস্যা হয়েছে। অ্যাডমিন ম্যানুয়ালি চেক করবেন।")
    
    del user_data[cid]

# ════════════════════════════════════════════════════════════════
#  CALLBACKS & ADMIN CONTROLS
# ════════════════════════════════════════════════════════════════
@bot.callback_query_handler(func=lambda c: True)
def callback_handler(call):
    cid  = call.message.chat.id
    uid  = call.from_user.id
    data = call.data

    if data == "cancel_flow":
        bot.answer_callback_query(call.id)
        user_data.pop(cid, None)
        bot.send_message(cid, "❌ বাতিল করা হয়েছে।", reply_markup=main_menu(cid))
        
    elif data in ["pay_bkash", "pay_nagad"]:
        bot.answer_callback_query(call.id)
        if cid not in user_data or user_data[cid].get("step") != "pay_method": return
        
        method_name = "𝐁𝐤𝐚𝐬𝐡" if data == "pay_bkash" else "𝐍𝐚𝐠𝐚𝐝"
        user_data[cid]["pay_method_name"] = method_name
        user_data[cid]["step"] = "number"
        
        msg = bot.send_message(cid, f"📱 *আপনার {method_name} নম্বর দিন:*\n_উদাহরণ: 01XXXXXXXXX_", parse_mode="Markdown", reply_markup=cancel_btn())
        bot.register_next_step_handler(msg, step_number)

    elif data.startswith("admrcv_"):
        if not is_admin(uid): 
            bot.answer_callback_query(call.id, "❌ 𝐀𝐜𝐜𝐞𝐬𝐬 𝐃𝐞𝐧𝐢𝐞𝐝!")
            return
        sub_id = data[7:]
        if sub_id in all_submissions:
            all_submissions[sub_id]["rcv_status"] = not all_submissions[sub_id].get("rcv_status", False)
            is_comp = all_submissions[sub_id].get("paid", False)
            live = all_submissions[sub_id].get("live_qty", 0)
            try: bot.edit_message_reply_markup(chat_id=cid, message_id=call.message.message_id, reply_markup=get_admin_channel_markup(sub_id, is_comp, live))
            except: pass
            bot.answer_callback_query(call.id, "✅ 𝐑𝐞𝐜𝐞𝐢𝐯𝐞 𝐒𝐭𝐚𝐭𝐮𝐬 𝐔𝐩𝐝𝐚𝐭𝐞𝐝!")

    elif data.startswith("admpay_"):
        if not is_admin(uid): 
            bot.answer_callback_query(call.id, "❌ 𝐀𝐜𝐜𝐞𝐬𝐬 𝐃𝐞𝐧𝐢𝐞𝐝!")
            return
        sub_id = data[7:]
        if sub_id in all_submissions:
            all_submissions[sub_id]["pay_status"] = not all_submissions[sub_id].get("pay_status", False)
            is_comp = all_submissions[sub_id].get("paid", False)
            live = all_submissions[sub_id].get("live_qty", 0)
            try: bot.edit_message_reply_markup(chat_id=cid, message_id=call.message.message_id, reply_markup=get_admin_channel_markup(sub_id, is_comp, live))
            except: pass
            bot.answer_callback_query(call.id, "✅ 𝐏𝐚𝐲𝐦𝐞𝐧𝐭 𝐒𝐭𝐚𝐭𝐮𝐬 𝐔𝐩𝐝𝐚𝐭𝐞𝐝!")

    elif data.startswith("review_"):
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        sub_id = data[7:]
        sub = all_submissions.get(sub_id)
        if not sub: bot.send_message(uid, "❌ 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 পাওয়া যায়নি।"); return
        if sub["paid"]: bot.send_message(uid, "✅ এটি আগেই 𝐏𝐚𝐢𝐝 করা হয়েছে।"); return

        ch_msg_id = call.message.message_id if cid == CHANNEL_ID else None

        msg = bot.send_message(uid, f"🔖 *𝐑𝐞𝐯𝐢𝐞𝐰: #{sub_id}*\n👤 𝐔𝐬𝐞𝐫: {sub['user']}\n📊 𝐓𝐨𝐭𝐚𝐥: {sub['qty']} 𝐩𝐜𝐬\n💳 𝐀𝐜𝐜𝐨𝐮𝐧𝐭: `{sub['account']}`\n\n✅ *কতটি 𝐋𝐢𝐯𝐞/𝐒𝐮𝐜𝐜𝐞𝐬𝐬? (সংখ্যা দিন):*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_review, sub_id, ch_msg_id)

    elif data == "admin_pending_subs":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        pending_list = {k: v for k, v in all_submissions.items() if not v.get("paid")}
        if not pending_list:
            bot.send_message(uid, "✅ কোনো 𝐏𝐞𝐧𝐝𝐢𝐧𝐠 সাবমিশন নেই!")
            return
        m = types.InlineKeyboardMarkup(row_width=1)
        for sid, s in list(pending_list.items())[:20]: 
            m.add(types.InlineKeyboardButton(f"⏳ {sid} | {s['user']} | {s['qty']} 𝐩𝐜𝐬", callback_data=f"review_{sid}"))
        m.add(types.InlineKeyboardButton("🔙 𝐁𝐚𝐜𝐤", callback_data="admin_refresh"))
        try: bot.edit_message_text("⏳ *𝐏𝐞𝐧𝐝𝐢𝐧𝐠 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧𝐬:*\n_রিভিউ করতে নিচের সাবমিশনগুলোতে ক্লিক করুন_", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data == "already_done":
        bot.answer_callback_query(call.id, "✅ 𝐀𝐥𝐫𝐞𝐚𝐝𝐲 𝐩𝐫𝐨𝐜𝐞𝐬𝐬𝐞𝐝!")

    elif data == "admin_msg_user":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        msg = bot.send_message(uid, "👤 ইউজারের 𝐓𝐞𝐥𝐞𝐠𝐫𝐚𝐦 𝐈𝐃 বা @𝐮𝐬𝐞𝐫𝐧𝐚𝐦𝐞 দিন:", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_msg_user_id)

    elif data == "admin_refresh":
        bot.answer_callback_query(call.id, "🔄 𝐑𝐞𝐟𝐫𝐞𝐬𝐡𝐞𝐝!")
        show_admin_panel(uid)

    elif data == "admin_all_subs":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        if not all_submissions: bot.send_message(uid, "📭 কোনো সাবমিশন নেই।"); return
        lines = ["📋 *𝐋𝐚𝐬𝐭 𝟐𝟎 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧𝐬:*", "━━━━━━━━━━━━━━━━━━"]
        for sid, s in list(all_submissions.items())[-20:]:
            icon = "✅" if s.get("paid") else "⏳"
            lines.append(f"{icon} `{sid}` | {s['user']} | ৳{s.get('total',0):.2f}")
        bot.send_message(uid, "\n".join(lines), parse_mode="Markdown")

    elif data == "admin_change_rate_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        m = types.InlineKeyboardMarkup(row_width=1)
        for code, cat in CATEGORIES.items():
            m.add(types.InlineKeyboardButton(f"✏️ {cat['name']}  →  {cat['rate']} 𝐓𝐤", callback_data=f"setrate_{code}"))
        m.add(types.InlineKeyboardButton("🔙 𝐁𝐚𝐜𝐤", callback_data="admin_refresh"))
        try: bot.edit_message_text("💰 *কোন 𝐂𝐚𝐭𝐞𝐠𝐨𝐫𝐲 এর 𝐑𝐚𝐭𝐞 পরিবর্তন করবেন?*", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data.startswith("setrate_"):
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        code = data[8:]
        if code not in CATEGORIES: return
        msg = bot.send_message(uid, f"✏️ *{CATEGORIES[code]['name']}*\nনতুন 𝐫𝐚𝐭𝐞 লিখুন (যেমন: 15.50):", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_set_rate, code)

    elif data == "admin_change_status_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        _show_status_menu(cid, call.message.message_id)

    elif data.startswith("changestatus_"):
        if not is_admin(uid): return
        code = data[13:]
        if code not in CATEGORIES: return
        cur = CATEGORIES[code]["status"]
        CATEGORIES[code]["status"] = "closed" if cur == "open" else "open"
        new = CATEGORIES[code]["status"]
        bot.answer_callback_query(call.id, f"{'✅ 𝐎𝐩𝐞𝐧' if new=='open' else '🔴 𝐂𝐥𝐨𝐬𝐞𝐝'}")
        _show_status_menu(cid, call.message.message_id)

    elif data == "admin_all_control":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        m = types.InlineKeyboardMarkup(row_width=1)
        m.add(types.InlineKeyboardButton("💳 𝐏𝐚𝐲𝐦𝐞𝐧𝐭 𝐌𝐞𝐭𝐡𝐨𝐝𝐬 𝐂𝐨𝐧𝐭𝐫𝐨𝐥", callback_data="admin_pay_control"))
        m.add(types.InlineKeyboardButton(f"📉 𝐌𝐢𝐧 𝐈𝐃 𝐋𝐢𝐦𝐢𝐭 (𝐍𝐨𝐰: {SYSTEM_SETTINGS['min_id_limit']})", callback_data="admin_change_min_limit"))
        m.add(types.InlineKeyboardButton(f"📝 𝐄𝐝𝐢𝐭 𝐖𝐞𝐥𝐜𝐨𝐦𝐞 𝐌𝐞𝐬𝐬𝐚𝐠𝐞", callback_data="admin_edit_welcome"))
        m.add(types.InlineKeyboardButton(f"📣 𝐒𝐞𝐭 𝐍𝐨𝐭𝐢𝐜𝐞 𝐁𝐨𝐚𝐫𝐝", callback_data="admin_set_notice"))
        
        maint_status = "🔴 𝐎𝐍 (𝐏𝐚𝐮𝐬𝐞𝐝)" if SYSTEM_SETTINGS["maintenance_mode"] else "✅ 𝐎𝐅𝐅 (𝐑𝐮𝐧𝐧𝐢𝐧𝐠)"
        m.add(types.InlineKeyboardButton(f"🛠️ 𝐌𝐚𝐢𝐧𝐭𝐞𝐧𝐚𝐧𝐜𝐞 𝐌𝐨𝐝𝐞: {maint_status}", callback_data="admin_toggle_maint"))
        m.add(types.InlineKeyboardButton("🚫 𝐁𝐚𝐧 / 𝐔𝐧𝐛𝐚𝐧 𝐔𝐬𝐞𝐫", callback_data="admin_ban_user_menu"))
        m.add(types.InlineKeyboardButton("💾 𝐁𝐚𝐜𝐤𝐮𝐩 𝐃𝐚𝐭𝐚 (𝐉𝐒𝐎𝐍)", callback_data="admin_db_backup"))
        m.add(types.InlineKeyboardButton("🔙 𝐁𝐚𝐜𝐤", callback_data="admin_refresh"))
        try:
            bot.edit_message_text("🎛️ *𝐒𝐲𝐬𝐭𝐞𝐦 𝐂𝐨𝐧𝐭𝐫𝐨𝐥 𝐏𝐚𝐧𝐞𝐥*", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data == "admin_pay_control":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        _show_payment_control(cid, call.message.message_id)
        
    elif data.startswith("toggle_pay_"):
        if not is_admin(uid): return
        method = data[11:] 
        key = f"pay_{method}"
        SYSTEM_SETTINGS[key] = not SYSTEM_SETTINGS.get(key, True)
        bot.answer_callback_query(call.id, f"✅ 𝐔𝐩𝐝𝐚𝐭𝐞𝐝!")
        _show_payment_control(cid, call.message.message_id)

    elif data == "admin_export_data":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        export_submissions_to_excel(uid)

    elif data == "admin_change_min_limit":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        msg = bot.send_message(uid, f"📉 *বর্তমান 𝐌𝐢𝐧𝐢𝐦𝐮𝐦 𝐈𝐃 𝐋𝐢𝐦𝐢𝐭:* {SYSTEM_SETTINGS['min_id_limit']} 𝐩𝐜𝐬\n\nনতুন 𝐋𝐢𝐦𝐢𝐭 কত দিতে চান? (সংখ্যায় লিখুন):", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_change_min_limit)

    elif data == "admin_edit_welcome":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        msg = bot.send_message(uid, "📝 *নতুন 𝐖𝐞𝐥𝐜𝐨𝐦𝐞 𝐌𝐞𝐬𝐬𝐚𝐠𝐞 লিখুন:*\n_(যে লেখাটি /𝐬𝐭𝐚𝐫𝐭 দিলে দেখাবে)_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_edit_welcome)
        
    elif data == "admin_ban_user_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        msg = bot.send_message(uid, "🚫 *ইউজারকে 𝐁𝐚𝐧 বা 𝐔𝐧𝐛𝐚𝐧 করতে তার 𝐓𝐞𝐥𝐞𝐠𝐫𝐚𝐦 𝐈𝐃 দিন:*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_ban_unban_user)

    elif data == "admin_db_backup":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        handle_db_backup(uid)

    elif data == "admin_toggle_maint":
        if not is_admin(uid): return
        SYSTEM_SETTINGS["maintenance_mode"] = not SYSTEM_SETTINGS["maintenance_mode"]
        bot.answer_callback_query(call.id, f"𝐌𝐚𝐢𝐧𝐭𝐞𝐧𝐚𝐧𝐜𝐞 𝐢𝐬 𝐧𝐨𝐰 {'𝐎𝐍' if SYSTEM_SETTINGS['maintenance_mode'] else '𝐎𝐅𝐅'}")
        call.data = "admin_all_control"
        callback_handler(call)

    elif data == "admin_set_notice":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        current = SYSTEM_SETTINGS.get("notice_board", "_(ফাঁকা)_")
        msg = bot.send_message(uid, f"📣 *𝐍𝐨𝐭𝐢𝐜𝐞 𝐁𝐨𝐚𝐫𝐝 আপডেট করুন*\n\nবর্তমান নোটিশ:\n{current}\n\nনতুন নোটিশ লিখুন:\n_(মুছে দিতে '𝐜𝐥𝐞𝐚𝐫' লিখুন)_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_set_notice)

    elif data == "admin_search_sub":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        msg = bot.send_message(uid, "🔎 *𝐒𝐮𝐛 𝐈𝐃 দিয়ে সার্চ করুন:*\n_উদাহরণ: 𝐒𝐔𝐁𝟎𝟎𝟎𝟏_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_admin_search_sub)

    elif data == "admin_user_history":
        bot.answer_callback_query(call.id)
        if not is_admin(uid): return
        msg = bot.send_message(uid, "👤 *যে ইউজারের 𝐇𝐢𝐬𝐭𝐨𝐫𝐲 দেখতে চান তার 𝐓𝐞𝐥𝐞𝐠𝐫𝐚𝐦 𝐈𝐃 দিন:*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_admin_view_user_history)

def _show_status_menu(cid, msg_id):
    m = types.InlineKeyboardMarkup(row_width=1)
    for code, cat in CATEGORIES.items():
        icon = "✅" if cat["status"] == "open" else "🔴"
        m.add(types.InlineKeyboardButton(f"{icon} {cat['name']}", callback_data=f"changestatus_{code}"))
    m.add(types.InlineKeyboardButton("🔙 𝐁𝐚𝐜𝐤", callback_data="admin_refresh"))
    try: bot.edit_message_text("⚙️ *𝐂𝐚𝐭𝐞𝐠𝐨𝐫𝐲 𝐎𝐍/𝐎𝐅𝐅 করুন:*\n_ক্লিক করলেই পরিবর্তন হবে_", cid, msg_id, reply_markup=m, parse_mode="Markdown")
    except: pass

def _show_payment_control(cid, msg_id):
    m = types.InlineKeyboardMarkup(row_width=1)
    b_status = "✅ 𝐎𝐍" if SYSTEM_SETTINGS.get("pay_bkash", True) else "🔴 𝐎𝐅𝐅"
    n_status = "✅ 𝐎𝐍" if SYSTEM_SETTINGS.get("pay_nagad", True) else "🔴 𝐎𝐅𝐅"
    
    m.add(types.InlineKeyboardButton(f"🟣 𝐁𝐤𝐚𝐬𝐡: {b_status}", callback_data="toggle_pay_bkash"))
    m.add(types.InlineKeyboardButton(f"🟠 𝐍𝐚𝐠𝐚𝐝: {n_status}", callback_data="toggle_pay_nagad"))
    m.add(types.InlineKeyboardButton("🔙 𝐁𝐚𝐜𝐤", callback_data="admin_all_control"))
    
    try: bot.edit_message_text("💳 *𝐏𝐚𝐲𝐦𝐞𝐧𝐭 𝐌𝐞𝐭𝐡𝐨𝐝𝐬 𝐂𝐨𝐧𝐭𝐫𝐨𝐥*\n_বাটনে ক্লিক করে অফ/অন করুন_", cid, msg_id, reply_markup=m, parse_mode="Markdown")
    except: pass

def export_submissions_to_excel(cid):
    if not all_submissions:
        bot.send_message(cid, "📭 এক্সপোর্ট করার মতো কোনো 𝐃𝐚𝐭𝐚 নেই।")
        return
    bot.send_message(cid, "⏳ 𝐄𝐱𝐩𝐨𝐫𝐭 𝐅𝐢𝐥𝐞 তৈরি হচ্ছে, অপেক্ষা করুন...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions Data"
        headers = ["Sub ID", "User Name", "Category", "Rate", "Submitted Qty", "Live/Approved Qty", "Total Amount (Tk)", "Account Number", "Note", "Date", "Status"]
        ws.append(headers)
        
        for sid, s in all_submissions.items():
            status = "Paid" if s.get("paid") else "Pending"
            ws.append([
                sid, s.get("user"), s.get("type_name"), s.get("rate"), s.get("qty"), 
                s.get("live_qty", 0), s.get("total", 0), s.get("account"), 
                s.get("note"), s.get("date"), status
            ])
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        filename = f"𝐄𝐱𝐩𝐨𝐫𝐭_𝐃𝐚𝐭𝐚_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        bot.send_document(cid, document=(filename, file_stream), caption="📊 *𝐀𝐥𝐥 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧𝐬 𝐄𝐱𝐩𝐨𝐫𝐭𝐞𝐝 𝐃𝐚𝐭𝐚*", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(cid, f"❌ এক্সপোর্ট করতে সমস্যা হয়েছে: {e}")

def step_review(message, sub_id, ch_msg_id):
    cid = message.chat.id
    if not (message.text or "").isdigit():
        msg = bot.send_message(cid, "❌ শুধু সংখ্যা দিন।")
        return bot.register_next_step_handler(msg, step_review, sub_id, ch_msg_id)

    live = int(message.text)
    sub  = all_submissions.get(sub_id)
    if not sub: return

    user_data[cid] = {
        "step": "review_screenshot",
        "review_sub_id": sub_id,
        "live_qty": live,
        "ch_msg_id": ch_msg_id
    }

    amount = live * sub["rate"]
    msg = bot.send_message(cid, f"✅ 𝐋𝐢𝐯𝐞: *{live}* টি।\n💰 𝐀𝐦𝐨𝐮𝐧𝐭: *৳{amount:.2f}*\n\n📸 *এবার 𝐏𝐚𝐲𝐦𝐞𝐧𝐭 এর স্ক্রিনশট পাঠান:*\n_(স্ক্রিনশট না দিতে চাইলে `𝐬𝐤𝐢𝐩` বা `না` লিখুন)_", parse_mode="Markdown")
    bot.register_next_step_handler(msg, step_review_screenshot)

def step_review_screenshot(message):
    cid = message.chat.id
    if cid not in user_data or "review_sub_id" not in user_data[cid]:
        return handle_text(message)

    d = user_data[cid]
    sub_id    = d["review_sub_id"]
    live      = d["live_qty"]
    ch_msg_id = d["ch_msg_id"]
    
    sub = all_submissions.get(sub_id)
    if not sub:
        del user_data[cid]
        return bot.send_message(cid, "❌ 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 পাওয়া যায়নি।")

    fail   = max(0, sub["qty"] - live)
    amount = live * sub["rate"]
    sub.update({"live_qty": live, "total": amount, "paid": True})

    result = (
        "📊 *𝐘𝐨𝐮𝐫 𝐑𝐞𝐬𝐮𝐥𝐭 𝐀𝐫𝐫𝐢𝐯𝐞𝐝!*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 𝐒𝐮𝐛 𝐈𝐃  : `{sub_id}`\n📁 𝐓𝐲𝐩𝐞    : {sub['type_name']}\n📅 𝐃𝐚𝐭𝐞    : {sub['date']}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 𝐓𝐨𝐭𝐚𝐥   : {sub['qty']} 𝐩𝐜𝐬\n✅ 𝐒𝐮𝐜𝐜𝐞𝐬𝐬 : *{live} 𝐩𝐜𝐬*\n❌ 𝐅𝐚𝐢𝐥𝐞𝐝  : *{fail} 𝐩𝐜𝐬*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💰 𝐀𝐦𝐨𝐮𝐧𝐭  : *{amount:.2f} 𝐓𝐤*\n💳 𝐀𝐜𝐜𝐨𝐮𝐧𝐭 : `{sub['account']}`\n\n"
        "✅ *𝐏𝐚𝐲𝐦𝐞𝐧𝐭 প্রক্রিয়া সম্পন্ন হয়েছে!* 🎉\n"
    )

    try: 
        if message.photo:
            bot.send_photo(sub["chat_id"], message.photo[-1].file_id, caption=result, parse_mode="Markdown")
        else:
            bot.send_message(sub["chat_id"], result, parse_mode="Markdown")
    except Exception as e: 
        log.warning(f"Result send failed: {e}")

    if ch_msg_id:
        done_btn = get_admin_channel_markup(sub_id, is_completed=True, live=live)
        try: bot.edit_message_reply_markup(CHANNEL_ID, ch_msg_id, reply_markup=done_btn)
        except: pass

    bot.send_message(cid, f"✅ *#{sub_id} 𝐑𝐞𝐬𝐮𝐥𝐭 পাঠানো হয়েছে!*\n𝐀𝐦𝐨𝐮𝐧𝐭: ৳{amount:.2f}", parse_mode="Markdown")
    del user_data[cid]

def step_set_notice(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    text = (message.text or "").strip()
    if text.lower() == "clear":
        SYSTEM_SETTINGS["notice_board"] = ""
        bot.send_message(cid, "✅ *𝐍𝐨𝐭𝐢𝐜𝐞 𝐁𝐨𝐚𝐫𝐝 মুছে দেওয়া হয়েছে!*", parse_mode="Markdown")
    else:
        SYSTEM_SETTINGS["notice_board"] = text
        bot.send_message(cid, "✅ *𝐍𝐨𝐭𝐢𝐜𝐞 𝐁𝐨𝐚𝐫𝐝 আপডেট করা হয়েছে!*\nইউজাররা এখন এটি দেখতে পাবে।", parse_mode="Markdown")
    show_admin_panel(cid)

def step_admin_search_sub(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    sub_id = (message.text or "").strip().upper()
    sub = all_submissions.get(sub_id)
    if not sub:
        bot.send_message(cid, f"❌ *'{sub_id}'* পাওয়া যায়নি।", parse_mode="Markdown")
        return
    
    status = "✅ 𝐏𝐚𝐢𝐝" if sub.get("paid") else "⏳ 𝐏𝐞𝐧𝐝𝐢𝐧𝐠"
    text = (
        f"🔎 *𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 𝐃𝐞𝐭𝐚𝐢𝐥𝐬: {sub_id}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 𝐔𝐬𝐞𝐫: {sub.get('user')} | `{sub.get('chat_id')}`\n"
        f"📁 𝐓𝐲𝐩𝐞: {sub.get('type_name')}\n"
        f"📅 𝐃𝐚𝐭𝐞: {sub.get('date')}\n"
        f"📦 𝐒𝐮𝐛𝐦𝐢𝐭: {sub.get('qty')} 𝐩𝐜𝐬\n"
        f"✅ 𝐋𝐢𝐯𝐞: {sub.get('live_qty', 0)} 𝐩𝐜𝐬\n"
        f"💰 𝐀𝐦𝐨𝐮𝐧𝐭: ৳{sub.get('total', 0):.2f}\n"
        f"💳 𝐀𝐜𝐜𝐨𝐮𝐧𝐭: `{sub.get('account')}`\n"
        f"📝 𝐍𝐨𝐭𝐞: {sub.get('note')}\n"
        f"📌 𝐒𝐭𝐚𝐭𝐮𝐬: {status}"
    )
    
    m = types.InlineKeyboardMarkup()
    if not sub.get("paid"):
        m.add(types.InlineKeyboardButton("🔍 𝐑𝐞𝐯𝐢𝐞𝐰 এখনই করুন", callback_data=f"review_{sub_id}"))
    bot.send_message(cid, text, parse_mode="Markdown", reply_markup=m)

def step_admin_view_user_history(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    
    target_input = (message.text or "").strip()
    target_id = None
    
    if target_input.lstrip("-").isdigit():
        target_id = int(target_input)
    elif target_input.lower() in username_to_id:
        target_id = username_to_id[target_input.lower()]
    
    if not target_id:
        bot.send_message(cid, "❌ ইউজার পাওয়া যায়নি।")
        return
    
    subs = user_submissions.get(target_id, [])
    if not subs:
        bot.send_message(cid, f"📭 এই ইউজারের (`{target_id}`) কোনো 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 নেই।", parse_mode="Markdown")
        return
    
    total_earned = 0.0
    lines = [f"📋 *𝐔𝐬𝐞𝐫 𝐇𝐢𝐬𝐭𝐨𝐫𝐲: `{target_id}`*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    for sid in reversed(subs[-20:]):
        s = all_submissions.get(sid, {})
        icon = "✅" if s.get("paid") else "⏳"
        amt = s.get("total", 0)
        total_earned += amt
        lines.append(f"{icon} `{sid}` | {s.get('type_name','?')} | ৳{amt:.2f}")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━")
    lines.append(f"💰 মোট আয়: *৳{total_earned:.2f}*")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

def step_edit_welcome(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    SYSTEM_SETTINGS["welcome_msg"] = message.text.strip()
    bot.send_message(cid, "✅ *𝐖𝐞𝐥𝐜𝐨𝐦𝐞 𝐌𝐞𝐬𝐬𝐚𝐠𝐞 সফলভাবে আপডেট করা হয়েছে!*", parse_mode="Markdown")
    show_admin_panel(cid)

def step_ban_unban_user(message):
    cid = message.chat.id
    target_id = message.text.strip()
    if not target_id.isdigit():
        bot.send_message(cid, "❌ শুধু ইউজারের 𝐓𝐞𝐥𝐞𝐠𝐫𝐚𝐦 𝐈𝐃 (সংখ্যায়) দিন।")
        return
    
    target_id = int(target_id)
    if target_id in ADMIN_IDS:
        bot.send_message(cid, "❌ 𝐀𝐝𝐦𝐢𝐧 কে 𝐁𝐚𝐧 করা যাবে না!")
        return

    if target_id in BANNED_USERS:
        BANNED_USERS.remove(target_id)
        bot.send_message(cid, f"✅ ইউজার `{target_id}` কে *𝐔𝐧𝐛𝐚𝐧* করা হয়েছে!", parse_mode="Markdown")
    else:
        BANNED_USERS.add(target_id)
        bot.send_message(cid, f"🚫 ইউজার `{target_id}` কে *𝐁𝐚𝐧* করা হয়েছে!", parse_mode="Markdown")

def handle_db_backup(cid):
    bot.send_message(cid, "⏳ 𝐃𝐚𝐭𝐚𝐛𝐚𝐬𝐞 𝐁𝐚𝐜𝐤𝐮𝐩 তৈরি হচ্ছে...")
    try:
        db_dump = {
            "submissions": all_submissions,
            "user_submissions": user_submissions,
            "registered_users": list(registered_users),
            "banned_users": list(BANNED_USERS),
            "settings": SYSTEM_SETTINGS
        }
        json_data = json.dumps(db_dump, indent=4, default=str)
        file_stream = BytesIO(json_data.encode("utf-8"))
        
        filename = f"𝐃𝐁_𝐁𝐚𝐜𝐤𝐮𝐩_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        bot.send_document(cid, document=(filename, file_stream), caption="💾 *𝐒𝐲𝐬𝐭𝐞𝐦 𝐃𝐚𝐭𝐚𝐛𝐚𝐬𝐞 𝐁𝐚𝐜𝐤𝐮𝐩*\n\nআপনার বটের সব সেভড 𝐃𝐚𝐭𝐚 এখানে আছে।", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(cid, f"❌ ব্যাকআপ নিতে সমস্যা হয়েছে: {e}")

def step_change_min_limit(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    try:
        new_limit = int(message.text.strip())
        SYSTEM_SETTINGS["min_id_limit"] = new_limit
        bot.send_message(cid, f"✅ *সফল!*\nএখন থেকে কেউ {new_limit} টির কম 𝐈𝐃 জমা দিতে পারবে না।", parse_mode="Markdown")
        show_admin_panel(cid)
    except ValueError:
        msg = bot.send_message(cid, "❌ ভুল ফরম্যাট! শুধু সংখ্যা দিন (যেমন: 10):")
        bot.register_next_step_handler(msg, step_change_min_limit)

def step_msg_user_id(message):
    uid_input = (message.text or "").strip()
    target = None
    if uid_input.lstrip("-").isdigit(): target = int(uid_input)
    elif uid_input.lower() in username_to_id: target = username_to_id[uid_input.lower()]
    else: bot.send_message(message.chat.id, "❌ *পাওয়া যায়নি!*", parse_mode="Markdown"); return
    msg = bot.send_message(message.chat.id, "✍️ 𝐌𝐞𝐬𝐬𝐚𝐠𝐞 বা ছবি পাঠান:")
    bot.register_next_step_handler(msg, step_send_to_user, target)

def step_send_to_user(message, target):
    try:
        if message.photo: bot.send_photo(target, message.photo[-1].file_id, caption=message.caption or "")
        else: bot.send_message(target, f"📩 *𝐀𝐝𝐦𝐢𝐧 𝐌𝐞𝐬𝐬𝐚𝐠𝐞:*\n\n{message.text}", parse_mode="Markdown")
        bot.send_message(message.chat.id, f"✅ পাঠানো হয়েছে → `{target}`", parse_mode="Markdown")
    except Exception as e: bot.send_message(message.chat.id, f"❌ ব্যর্থ: {e}")

def step_broadcast(message):
    if message.text in MENU_BUTTONS: return handle_text(message)
    ok = fail = 0
    bot.send_message(message.chat.id, f"⏳ 𝐁𝐫𝐨𝐚𝐝𝐜𝐚𝐬𝐭 শুরু...")
    for uid in list(registered_users):
        try:
            if message.photo: 
                bot.send_photo(uid, message.photo[-1].file_id, caption=message.caption or "")
            elif message.sticker:
                bot.send_sticker(uid, message.sticker.file_id)
            else: 
                bot.send_message(uid, message.text, parse_mode="Markdown")
            ok += 1
        except: fail += 1
        time.sleep(0.05)
    bot.send_message(message.chat.id, f"📢 *𝐁𝐫𝐨𝐚𝐝𝐜𝐚𝐬𝐭 𝐒𝐮𝐜𝐜𝐞𝐬𝐬𝐟𝐮𝐥!*\n✅ সফল: {ok}\n❌ ব্যর্থ: {fail}", parse_mode="Markdown")

def step_set_rate(message, code):
    try:
        new_rate = float(message.text.strip())
        CATEGORIES[code]["rate"] = new_rate
        bot.send_message(message.chat.id, f"✅ *{CATEGORIES[code]['name']}*\nনতুন 𝐫𝐚𝐭𝐞: *{new_rate:.2f} 𝐓𝐤*", parse_mode="Markdown")
        show_admin_panel(message.chat.id)
    except ValueError:
        msg = bot.send_message(message.chat.id, "❌ সংখ্যা দিন:")
        bot.register_next_step_handler(msg, step_set_rate, code)

def show_history(cid):
    subs = user_submissions.get(cid, [])
    if not subs: bot.send_message(cid, "📭 এখনো কোনো 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 নেই।"); return
    lines = ["📜 *𝐌𝐲 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧 𝐇𝐢𝐬𝐭𝐨𝐫𝐲*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    for sid in reversed(subs[-15:]):
        s    = all_submissions.get(sid, {})
        icon = "✅" if s.get("paid") else "⏳"
        amt  = s.get("total", 0)
        lines.append(f"{icon} `{sid}` | {s.get('type_name','?')} | ৳{amt:.2f} | {s.get('date','?')}")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

def show_admin_panel(cid):
    total  = len(all_submissions)
    paid   = sum(1 for s in all_submissions.values() if s.get("paid"))
    tk     = sum(s.get("total", 0) for s in all_submissions.values() if s.get("paid"))
    users  = len(registered_users)
    banned = len(BANNED_USERS)
    
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(types.InlineKeyboardButton("⏳ 𝐏𝐞𝐧𝐝𝐢𝐧𝐠", callback_data="admin_pending_subs"),
          types.InlineKeyboardButton("📋 𝐀𝐥𝐥 𝐒𝐮𝐛𝐬", callback_data="admin_all_subs"))
    m.add(types.InlineKeyboardButton("💰 𝐑𝐚𝐭𝐞𝐬", callback_data="admin_change_rate_menu"),
          types.InlineKeyboardButton("⚙️ 𝐂𝐚𝐭𝐞𝐠𝐨𝐫𝐢𝐞𝐬", callback_data="admin_change_status_menu"))
    m.add(types.InlineKeyboardButton("👁️ 𝐔𝐬𝐞𝐫 𝐇𝐢𝐬𝐭𝐨𝐫𝐲", callback_data="admin_user_history"),
          types.InlineKeyboardButton("📩 𝐌𝐬𝐠 𝐔𝐬𝐞𝐫", callback_data="admin_msg_user"))
    m.add(types.InlineKeyboardButton("🔎 𝐒𝐞𝐚𝐫𝐜𝐡 𝐈𝐃", callback_data="admin_search_sub"),
          types.InlineKeyboardButton("📥 𝐁𝐚𝐜𝐤𝐮𝐩", callback_data="admin_export_data"))
    m.add(types.InlineKeyboardButton("🎛️ 𝐒𝐲𝐬𝐭𝐞𝐦 𝐒𝐞𝐭𝐭𝐢𝐧𝐠𝐬", callback_data="admin_all_control"),
          types.InlineKeyboardButton("🔄 𝐑𝐞𝐟𝐫𝐞𝐬𝐡", callback_data="admin_refresh"))
    
    admin_text = (
        "👑 *𝐀𝐃𝐕𝐀𝐍𝐂𝐄𝐃 𝐀𝐃𝐌𝐈𝐍 𝐂𝐎𝐍𝐓𝐑𝐎𝐋*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 𝐓𝐨𝐭𝐚𝐥 𝐔𝐬𝐞𝐫𝐬 : *{users}*\n"
        f"🚫 𝐁𝐚𝐧𝐧𝐞𝐝      : *{banned}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📥 𝐒𝐮𝐛𝐦𝐢𝐬𝐬𝐢𝐨𝐧𝐬 : *{total}*\n"
        f"✅ 𝐏𝐚𝐢𝐝        : *{paid}*\n"
        f"⏳ 𝐏𝐞𝐧𝐝𝐢𝐧𝐠     : *{total - paid}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💵 𝐓𝐨𝐭𝐚𝐥 𝐏𝐚𝐢𝐝  : *৳{tk:.2f}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    bot.send_message(cid, admin_text, reply_markup=m, parse_mode="Markdown")

@bot.message_handler(func=lambda m: True, content_types=["photo", "video", "audio", "voice", "sticker"])
def handle_media(message):
    cid = message.chat.id
    if is_banned(cid): return
    if cid in user_data and user_data[cid].get("step") == "file":
        bot.send_message(cid, "❌ একটি *.𝐱𝐥𝐬𝐱* ফাইল পাঠান।", reply_markup=cancel_btn(), parse_mode="Markdown")
    else:
        bot.send_message(cid, "❓ শুধু নিচের বাটনগুলো ব্যবহার করুন।", reply_markup=main_menu(cid))

# ════════════════════════════════════════════════════════════════
#  RUN
# ════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    log.info(f"🚀 {BOT_NAME} starting...")
    while True:
        try:
            bot.infinity_polling(timeout=60, long_polling_timeout=30, skip_pending=True)
        except Exception as e:
            log.error(f"Polling crashed: {e}")
            time.sleep(5)
bot = telebot.TeleBot(BOT_TOKEN, parse_mode=None)

user_data        = {}
all_submissions  = {}
user_submissions = {}
submission_count = 0
registered_users = set()
BANNED_USERS     = set()
username_to_id   = {}

# ════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════
def is_admin(cid):        
    return cid in ADMIN_IDS

def is_banned(cid):
    return cid in BANNED_USERS

def safe_md(text):
    escape_chars = r'_*[]()~`>#+-=|{}.!'
    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', str(text))

MENU_BUTTONS = ["🚀 SELL ID 🚀", "📊 Price List", "👑 Owner Support", "👨‍💻 Admin Support", "📜 My History", "👤 My Profile", "📣 Notice Board", "🔐 Admin Panel", "📢 Broadcast"]

def main_menu(cid):
    m = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    m.add(types.KeyboardButton("🚀 SELL ID 🚀"))
    m.row(types.KeyboardButton("📊 Price List"), types.KeyboardButton("📜 My History"))
    m.row(types.KeyboardButton("👑 Owner Support"), types.KeyboardButton("👨‍💻 Admin Support"))
    m.row(types.KeyboardButton("👤 My Profile"), types.KeyboardButton("📣 Notice Board"))
    
    if is_admin(cid):
        m.row(types.KeyboardButton("🔐 Admin Panel"), types.KeyboardButton("📢 Broadcast"))
    return m

def cancel_btn():
    m = types.InlineKeyboardMarkup()
    m.add(types.InlineKeyboardButton("❌ CANCEL", callback_data="cancel_flow"))
    return m

def build_price_list():
    lines = [
        "🏷 *Price List*",
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    ]
    for cat in CATEGORIES.values():
        st = "✅ Open" if cat["status"] == "open" else "🔴 Closed"
        lines.append(
            f"\n{'✅' if cat['status']=='open' else '🔴'} *{cat['name']}*\n"
            f"   💰 Rate: *{cat['rate']:.2f} Tk/pcs*\n"
            f"   📋 `{cat['format']}`\n"
            f"   Status: {st}"
        )
    lines.append("\n━━━━━━━━━━━━━━━━━━━━━━━━")
    return "\n".join(lines)

def register_user(message):
    registered_users.add(message.chat.id)
    if message.from_user and message.from_user.username:
        username_to_id[f"@{message.from_user.username.lower()}"] = message.chat.id

# ════════════════════════════════════════════════════════════════
#  /start
# ════════════════════════════════════════════════════════════════
@bot.message_handler(commands=["start"])
def cmd_start(message):
    cid = message.chat.id
    if is_banned(cid): return

    if SYSTEM_SETTINGS["maintenance_mode"] and not is_admin(cid):
        bot.send_message(cid, "🛠️ সিস্টেম বর্তমানে আপডেটের কাজে সাময়িকভাবে বন্ধ আছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।")
        return

    register_user(message)
    name = message.from_user.first_name or "বন্ধু"

    welcome = (
        f"👋 আসসালামু আলাইকুম, *{name}*!\n\n"
        f"🔥 *{BOT_NAME}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"{SYSTEM_SETTINGS['welcome_msg']}\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    bot.send_message(cid, welcome, reply_markup=main_menu(cid), parse_mode="Markdown")

# ════════════════════════════════════════════════════════════════
#  MAIN TEXT HANDLER
# ════════════════════════════════════════════════════════════════
@bot.message_handler(func=lambda m: True, content_types=["text"])
def handle_text(message):
    cid  = message.chat.id
    
    if is_banned(cid): return
    
    if SYSTEM_SETTINGS["maintenance_mode"] and not is_admin(cid):
        bot.send_message(cid, "🛠️ সিস্টেম বর্তমানে আপডেটের কাজে সাময়িকভাবে বন্ধ আছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।")
        return

    register_user(message)
    text = (message.text or "").strip()

    if text == "🚀 SELL ID 🚀":
        show_categories(cid)

    elif text == "📊 Price List":
        bot.send_message(cid, build_price_list(), parse_mode="Markdown")

    elif text == "👑 Owner Support":
        bot.send_message(cid, f"👑 *অনার সাপোর্ট:*\n\n👤 {OWNER_SUPPORT}\n\n_যেকোনো জরুরি সমস্যায় সরাসরি মেসেজ করুন।_", parse_mode="Markdown")

    elif text == "👨‍💻 Admin Support":
        bot.send_message(cid, f"👨‍💻 *এডমিন সাপোর্ট:*\n\n👤 {ADMIN_SUPPORT}\n\n_সাধারণ সমস্যা বা আইডি সেলের বিষয়ে মেসেজ করুন।_", parse_mode="Markdown")

    elif text == "📜 My History":
        show_history(cid)

    elif text == "👤 My Profile":
        show_profile(cid)

    elif text == "📣 Notice Board":
        notice = SYSTEM_SETTINGS.get("notice_board", "")
        if not notice:
            bot.send_message(cid, "📣 *Notice Board*\n━━━━━━━━━━━━━━━━━━━━━━━━\n_এখনো কোনো নোটিশ নেই।_", parse_mode="Markdown")
        else:
            bot.send_message(cid, f"📣 *Notice Board*\n━━━━━━━━━━━━━━━━━━━━━━━━\n{notice}", parse_mode="Markdown")

    elif text == "🔐 Admin Panel":
        if is_admin(cid):
            show_admin_panel(cid)
            
    elif text == "📢 Broadcast":
        if is_admin(cid):
            msg = bot.send_message(cid, "📢 *ব্রডকাস্ট মেসেজ লিখুন*\n(আপনি টেক্সট, ছবি বা স্টিকারও পাঠাতে পারেন):", parse_mode="Markdown", reply_markup=cancel_btn())
            bot.register_next_step_handler(msg, step_broadcast)

    else:
        bot.send_message(cid, "❓ বুঝতে পারিনি। নিচের বাটনগুলো ব্যবহার করুন।", reply_markup=main_menu(cid))

def show_profile(cid):
    subs = user_submissions.get(cid, [])
    total_submitted = len(subs)
    total_approved = 0
    total_earned = 0.0

    for sid in subs:
        s = all_submissions.get(sid, {})
        if s.get("paid"):
            total_approved += s.get("live_qty", 0)
            total_earned += s.get("total", 0.0)

    profile_text = (
        "👤 *আপনার প্রোফাইল*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 User ID: `{cid}`\n"
        f"📦 মোট সাবমিট: *{total_submitted}* বার\n"
        f"✅ মোট এপ্রুভড আইডি: *{total_approved}* টি\n"
        f"💰 মোট আয়: *৳{total_earned:.2f}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    bot.send_message(cid, profile_text, parse_mode="Markdown")

# ════════════════════════════════════════════════════════════════
#  SELL FLOW
# ════════════════════════════════════════════════════════════════
def show_categories(cid):
    markup = types.InlineKeyboardMarkup(row_width=1)
    for code, cat in CATEGORIES.items():
        if cat["status"] == "open":
            label = f"✅ {cat['name']}  —  {cat['rate']:.2f} Tk"
        else:
            label = f"🔴 {cat['name']}  [বন্ধ]"
        markup.add(types.InlineKeyboardButton(label, callback_data=f"sell_{code}"))

    bot.send_message(cid, "🛒 *ID বিক্রি করুন*\n━━━━━━━━━━━━━━━━━━━━━━━━\n\n👇 *ক্যাটাগরি সিলেক্ট করুন:*", reply_markup=markup, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("sell_"))
def cb_sell(call):
    bot.answer_callback_query(call.id)
    cid  = call.message.chat.id
    code = call.data[5:]

    if code not in CATEGORIES: return
    cat = CATEGORIES[code]
    if cat["status"] != "open":
        bot.send_message(cid, "🔴 এই ক্যাটাগরি বর্তমানে বন্ধ আছে।")
        return

    user_data[cid] = {
        "type": code, "type_name": cat["name"],
        "rate": cat["rate"], "format": cat["format"],
        "date": datetime.now().strftime("%d %b %Y"),
        "step": "file" # Changed from username to file
    }

    msg = bot.send_message(
        cid,
        f"✅ *সিলেক্ট:* {cat['name']}\n💰 Rate: *{cat['rate']:.2f} Tk/pcs*\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📤 *এখন Excel ফাইল পাঠান*\n📋 Column Layout:\n`{cat['format']}`\n\n"
        f"⚠️ শুধুমাত্র *.xlsx* ফাইল সাপোর্ট করে",
        reply_markup=cancel_btn(), parse_mode="Markdown"
    )
    bot.register_next_step_handler(msg, step_file_text)

def step_file_text(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if message.content_type == "document": return handle_docs(message)
    bot.send_message(cid, "❌ একটি *.xlsx* ফাইল পাঠান।", reply_markup=cancel_btn(), parse_mode="Markdown")
    bot.register_next_step_handler(message, step_file_text)

@bot.message_handler(content_types=["document"])
def handle_docs(message):
    cid = message.chat.id
    if is_banned(cid): return
    if cid not in user_data or user_data[cid].get("step") != "file": return

    fname = message.document.file_name or ""
    if not fname.lower().endswith(".xlsx"):
        msg = bot.reply_to(message, "❌ শুধু *.xlsx* ফাইল সাপোর্ট করে!\nএকটি Excel ফাইল পাঠান:", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_file_text)

    try:
        fi   = bot.get_file(message.document.file_id)
        raw  = bot.download_file(fi.file_path)
        wb   = openpyxl.load_workbook(filename=BytesIO(raw), data_only=True)
        ws   = wb.active
        
        uids = []
        valid_rows = 0
        for row in ws.iter_rows(values_only=True):
            # Check if row is not completely empty
            if any(c is not None and str(c).strip() for c in row):
                uid = str(row[0]).strip() if row and row[0] is not None else ""
                if uid: 
                    uids.append(uid)
                valid_rows += 1

        qty = valid_rows
        min_limit = SYSTEM_SETTINGS["min_id_limit"]
        
        if qty < min_limit:
            msg = bot.reply_to(message, f"❌ আপনি মাত্র *{qty}* টি আইডি দিয়েছেন।\n⚠️ মিনিমাম *{min_limit}* টি আইডি একসাথে দিতে হবে!\n\nসঠিক ফাইলটি আবার পাঠান:", reply_markup=cancel_btn(), parse_mode="Markdown")
            return bot.register_next_step_handler(msg, step_file_text)

        # -------------------------------------------------------------
        #  VALIDATION 1: Duplicate ID Check
        # -------------------------------------------------------------
        if len(uids) != len(set(uids)):
            msg = bot.reply_to(message, "❌ *ফাইলে ডুপ্লিকেট (একই আইডি একাধিকবার) পাওয়া গেছে!*\nঅনুগ্রহ করে ডুপ্লিকেট রিমুভ করে সম্পূর্ণ ফ্রেশ ফাইলটি আবার দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
            return bot.register_next_step_handler(msg, step_file_text)
            
        # -------------------------------------------------------------
        #  VALIDATION 2: Category Prefix Check
        # -------------------------------------------------------------
        category_code = user_data[cid]["type"]
        if category_code == "PC1000X" and any(not str(u).startswith("1000") for u in uids):
            msg = bot.reply_to(message, "❌ *ভুল ক্যাটাগরি!*\nআপনি 1000x সিলেক্ট করেছেন কিন্তু ফাইলে অন্য আইডি (যেমন 61x) পাওয়া গেছে। সঠিক ফাইল দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
            return bot.register_next_step_handler(msg, step_file_text)
            
        elif category_code == "PC6155X" and any(not (str(u).startswith("6155") or str(u).startswith("6156") or str(u).startswith("6157")) for u in uids):
             msg = bot.reply_to(message, "❌ *ভুল ক্যাটাগরি!*\nআপনি 6155x/56x/57x সিলেক্ট করেছেন কিন্তু ফাইলে অন্য আইডি পাওয়া গেছে। সঠিক ফাইল দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
             return bot.register_next_step_handler(msg, step_file_text)
             
        elif category_code == "PC6158X" and any(not str(u).startswith("6158") for u in uids):
             msg = bot.reply_to(message, "❌ *ভুল ক্যাটাগরি!*\nআপনি 6158x সিলেক্ট করেছেন কিন্তু ফাইলে অন্য আইডি পাওয়া গেছে। সঠিক ফাইল দিন:", reply_markup=cancel_btn(), parse_mode="Markdown")
             return bot.register_next_step_handler(msg, step_file_text)

        # Validation Passed!
        user_data[cid].update({"qty": qty, "file_name": fname, "file_id": message.document.file_id, "step": "username"})

        msg = bot.reply_to(
            message,
            f"✅ *ফাইল গৃহীত হয়েছে! (No Duplicates)*\n📊 মোট row: *{qty} pcs*\n\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "👤 *আপনার Telegram username দিন:*\n_উদাহরণ: @myusername_",
            reply_markup=cancel_btn(), parse_mode="Markdown"
        )
        bot.register_next_step_handler(msg, step_username)

    except Exception as e:
        log.error(f"File error: {e}")
        msg = bot.reply_to(message, f"❌ ফাইল পড়তে সমস্যা: `{e}`\nআবার চেষ্টা করুন:", reply_markup=cancel_btn(), parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_file_text)


def step_username(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    uname = (message.text or "").strip()
    if not re.match(r'^@[a-zA-Z0-9_]{4,32}$', uname):
        msg = bot.send_message(cid, "❌ *ভুল ফরম্যাট!*\n@ দিয়ে শুরু করুন\n_উদাহরণ: @myusername_", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_username)

    username_to_id[uname.lower()] = cid
    user_data[cid]["user"] = uname
    user_data[cid]["step"] = "pay_method"

    # Check available payment methods
    m = types.InlineKeyboardMarkup(row_width=2)
    btns = []
    if SYSTEM_SETTINGS.get("pay_bkash", True):
        btns.append(types.InlineKeyboardButton("🟣 Bkash", callback_data="pay_bkash"))
    if SYSTEM_SETTINGS.get("pay_nagad", True):
        btns.append(types.InlineKeyboardButton("🟠 Nagad", callback_data="pay_nagad"))

    if not btns:
        bot.send_message(cid, "❌ বর্তমানে পেমেন্ট গ্রহণ বন্ধ আছে। অ্যাডমিনকে জানান।", reply_markup=main_menu(cid))
        del user_data[cid]
        return

    m.add(*btns)
    m.add(types.InlineKeyboardButton("❌ CANCEL", callback_data="cancel_flow"))

    bot.send_message(
        cid,
        "💳 *পেমেন্ট মেথড সিলেক্ট করুন:*",
        reply_markup=m, parse_mode="Markdown"
    )

def step_number(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    number = (message.text or "").strip()
    if not re.match(r'^01[0-9]{9}$', number):
        msg = bot.send_message(cid, "❌ *ভুল নম্বর!* ১১ ডিজিটের বাংলাদেশী নম্বর দিন\n_উদাহরণ: 01712345678_", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_number)

    method_name = user_data[cid].get("pay_method_name", "Unknown")
    user_data[cid]["account"] = f"{method_name} - {number}"
    user_data[cid]["step"]    = "note"

    msg = bot.send_message(cid, "📝 *কোনো নোট আছে?*\n_না থাকলে শুধু 'না' লিখুন_", reply_markup=cancel_btn(), parse_mode="Markdown")
    bot.register_next_step_handler(msg, step_note)

def step_note(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    note = (message.text or "").strip()
    finalize(cid, note)

def finalize(cid, note):
    global submission_count
    submission_count += 1
    sub_id = f"SUB{submission_count:04d}"
    d      = user_data[cid]
    est    = d["qty"] * d["rate"]

    all_submissions[sub_id] = {
        "chat_id":   cid,
        "user":      d["user"],
        "type_name": d["type_name"],
        "rate":      d["rate"],
        "qty":       d["qty"],
        "total":     0,
        "account":   d["account"],
        "note":      note,
        "date":      d["date"],
        "paid":      False,
        "live_qty":  0,
        "file_name": d["file_name"],
    }
    user_submissions.setdefault(cid, []).append(sub_id)

    receipt = (
        "✅ *সাবমিশন সম্পন্ন হয়েছে!*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 Sub ID  : `{sub_id}`\n📁 Type    : {d['type_name']}\n📅 Date    : {d['date']}\n📄 File    : `{d['file_name']}`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 Total   : *{d['qty']} pcs*\n💰 Est.Amt : *{est:.2f} Tk* ({d['rate']} × {d['qty']})\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💳 Account : `{d['account']}`\n📝 Note    : {note}\n\n"
        "⏳ *Admin review করার পর result পাঠাবে।*"
    )

    adm_cap = (
        f"📥 *New Submission #{sub_id}*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 User    : {safe_md(d['user'])} | `{cid}`\n🏷️ Type    : {safe_md(d['type_name'])}\n📊 Rows    : *{d['qty']} pcs*\n"
        f"💰 Est.    : *{est:.2f} Tk*\n💳 Account : `{safe_md(d['account'])}`\n📝 Note    : {safe_md(note)}"
    )
    adm_btn = types.InlineKeyboardMarkup()
    adm_btn.add(types.InlineKeyboardButton("🟢 Review / Result দিন", callback_data=f"review_{sub_id}"))

    bot.send_message(cid, receipt, parse_mode="Markdown", reply_markup=main_menu(cid))
    try:
        bot.send_document(CHANNEL_ID, d["file_id"], caption=adm_cap, parse_mode="Markdown", reply_markup=adm_btn)
    except Exception as e: 
        log.error(f"Channel send error: {e}")
        bot.send_message(cid, "⚠️ ফাইল জমা হয়েছে কিন্তু চ্যানেলে ফরোয়ার্ড হতে সমস্যা হয়েছে। অ্যাডমিন ম্যানুয়ালি চেক করবেন।")
    
    del user_data[cid]

# ════════════════════════════════════════════════════════════════
#  CALLBACKS & ADMIN CONTROLS
# ════════════════════════════════════════════════════════════════
@bot.callback_query_handler(func=lambda c: True)
def callback_handler(call):
    cid  = call.message.chat.id
    data = call.data

    if data == "cancel_flow":
        bot.answer_callback_query(call.id)
        user_data.pop(cid, None)
        bot.send_message(cid, "❌ বাতিল করা হয়েছে।", reply_markup=main_menu(cid))
        
    elif data in ["pay_bkash", "pay_nagad"]:
        bot.answer_callback_query(call.id)
        if cid not in user_data or user_data[cid].get("step") != "pay_method": return
        
        method_name = "Bkash" if data == "pay_bkash" else "Nagad"
        user_data[cid]["pay_method_name"] = method_name
        user_data[cid]["step"] = "number"
        
        msg = bot.send_message(cid, f"📱 *আপনার {method_name} নম্বর দিন:*\n_উদাহরণ: 01XXXXXXXXX_", parse_mode="Markdown", reply_markup=cancel_btn())
        bot.register_next_step_handler(msg, step_number)

    elif data.startswith("review_"):
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        sub_id = data[7:]
        sub = all_submissions.get(sub_id)
        if not sub: bot.send_message(cid, "❌ Submission পাওয়া যায়নি।"); return
        if sub["paid"]: bot.send_message(cid, "✅ এটি আগেই পেইড করা হয়েছে।"); return

        ch_msg_id = call.message.message_id if call.message.chat.id == CHANNEL_ID else None

        msg = bot.send_message(cid, f"🔖 *Review: #{sub_id}*\n👤 User: {sub['user']}\n📊 Total: {sub['qty']} pcs\n💳 Account: `{sub['account']}`\n\n✅ *কতটি Live/Success? (সংখ্যা দিন):*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_review, sub_id, ch_msg_id)

    elif data == "admin_pending_subs":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        pending_list = {k: v for k, v in all_submissions.items() if not v.get("paid")}
        if not pending_list:
            bot.send_message(cid, "✅ কোনো পেন্ডিং সাবমিশন নেই!")
            return
        m = types.InlineKeyboardMarkup(row_width=1)
        for sid, s in list(pending_list.items())[:20]: 
            m.add(types.InlineKeyboardButton(f"⏳ {sid} | {s['user']} | {s['qty']} pcs", callback_data=f"review_{sid}"))
        m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
        try: bot.edit_message_text("⏳ *পেন্ডিং সাবমিশন তালিকা:*\n_রিভিউ করতে নিচের সাবমিশনগুলোতে ক্লিক করুন_", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data == "already_done":
        bot.answer_callback_query(call.id, "✅ Already processed!")

    elif data == "admin_msg_user":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "👤 ইউজারের Telegram ID বা @username দিন:", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_msg_user_id)

    elif data == "admin_refresh":
        bot.answer_callback_query(call.id, "🔄 Refreshed!")
        show_admin_panel(cid)

    elif data == "admin_all_subs":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        if not all_submissions: bot.send_message(cid, "📭 কোনো সাবমিশন নেই।"); return
        lines = ["📋 *সর্বশেষ ২০ সাবমিশন:*", "━━━━━━━━━━━━━━━━━━"]
        for sid, s in list(all_submissions.items())[-20:]:
            icon = "✅" if s.get("paid") else "⏳"
            lines.append(f"{icon} `{sid}` | {s['user']} | ৳{s.get('total',0):.2f}")
        bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

    elif data == "admin_change_rate_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        m = types.InlineKeyboardMarkup(row_width=1)
        for code, cat in CATEGORIES.items():
            m.add(types.InlineKeyboardButton(f"✏️ {cat['name']}  →  {cat['rate']} Tk", callback_data=f"setrate_{code}"))
        m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
        try: bot.edit_message_text("💰 *কোন ক্যাটাগরির রেট পরিবর্তন করবেন?*", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data.startswith("setrate_"):
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        code = data[8:]
        if code not in CATEGORIES: return
        msg = bot.send_message(cid, f"✏️ *{CATEGORIES[code]['name']}*\nনতুন rate লিখুন (যেমন: 15.50):", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_set_rate, code)

    elif data == "admin_change_status_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        _show_status_menu(cid, call.message.message_id)

    elif data.startswith("changestatus_"):
        if not is_admin(cid): return
        code = data[13:]
        if code not in CATEGORIES: return
        cur = CATEGORIES[code]["status"]
        CATEGORIES[code]["status"] = "closed" if cur == "open" else "open"
        new = CATEGORIES[code]["status"]
        bot.answer_callback_query(call.id, f"{'✅ Open' if new=='open' else '🔴 Closed'}")
        _show_status_menu(cid, call.message.message_id)

    elif data == "admin_all_control":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        m = types.InlineKeyboardMarkup(row_width=1)
        m.add(types.InlineKeyboardButton("💳 Payment Methods Control", callback_data="admin_pay_control"))
        m.add(types.InlineKeyboardButton(f"📉 Min ID Limit (Now: {SYSTEM_SETTINGS['min_id_limit']})", callback_data="admin_change_min_limit"))
        m.add(types.InlineKeyboardButton(f"📝 Edit Welcome Message", callback_data="admin_edit_welcome"))
        m.add(types.InlineKeyboardButton(f"📣 Set Notice Board", callback_data="admin_set_notice"))
        
        maint_status = "🔴 ON (Paused)" if SYSTEM_SETTINGS["maintenance_mode"] else "✅ OFF (Running)"
        m.add(types.InlineKeyboardButton(f"🛠️ Maintenance Mode: {maint_status}", callback_data="admin_toggle_maint"))
        m.add(types.InlineKeyboardButton("🚫 Ban / Unban User", callback_data="admin_ban_user_menu"))
        m.add(types.InlineKeyboardButton("💾 Backup Data (JSON)", callback_data="admin_db_backup"))
        m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
        try:
            bot.edit_message_text("🎛️ *System Control Panel*", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data == "admin_pay_control":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        _show_payment_control(cid, call.message.message_id)
        
    elif data.startswith("toggle_pay_"):
        if not is_admin(cid): return
        method = data[11:] # 'bkash' or 'nagad'
        key = f"pay_{method}"
        SYSTEM_SETTINGS[key] = not SYSTEM_SETTINGS.get(key, True)
        bot.answer_callback_query(call.id, f"✅ Updated!")
        _show_payment_control(cid, call.message.message_id)

    elif data == "admin_export_data":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        export_submissions_to_excel(cid)

    elif data == "admin_change_min_limit":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, f"📉 *বর্তমান মিনিমাম আইডি লিমিট:* {SYSTEM_SETTINGS['min_id_limit']} pcs\n\nনতুন মিনিমাম লিমিট কত দিতে চান? (সংখ্যায় লিখুন):", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_change_min_limit)

    elif data == "admin_edit_welcome":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "📝 *নতুন ওয়েলকাম মেসেজ লিখুন:*\n_(যে লেখাটি /start দিলে দেখাবে)_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_edit_welcome)
        
    elif data == "admin_ban_user_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "🚫 *ইউজারকে ব্যান বা আনব্যান করতে তার Telegram ID দিন:*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_ban_unban_user)

    elif data == "admin_db_backup":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        handle_db_backup(cid)

    elif data == "admin_toggle_maint":
        if not is_admin(cid): return
        SYSTEM_SETTINGS["maintenance_mode"] = not SYSTEM_SETTINGS["maintenance_mode"]
        bot.answer_callback_query(call.id, f"Maintenance is now {'ON' if SYSTEM_SETTINGS['maintenance_mode'] else 'OFF'}")
        call.data = "admin_all_control"
        callback_handler(call)

    elif data == "admin_set_notice":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        current = SYSTEM_SETTINGS.get("notice_board", "_(ফাঁকা)_")
        msg = bot.send_message(cid, f"📣 *Notice Board আপডেট করুন*\n\nবর্তমান নোটিশ:\n{current}\n\nনতুন নোটিশ লিখুন:\n_(মুছে দিতে 'clear' লিখুন)_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_set_notice)

    elif data == "admin_search_sub":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "🔎 *Sub ID দিয়ে সার্চ করুন:*\n_উদাহরণ: SUB0001_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_admin_search_sub)

    elif data == "admin_user_history":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "👤 *যে ইউজারের হিস্টোরি দেখতে চান তার Telegram ID দিন:*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_admin_view_user_history)

def _show_status_menu(cid, msg_id):
    m = types.InlineKeyboardMarkup(row_width=1)
    for code, cat in CATEGORIES.items():
        icon = "✅" if cat["status"] == "open" else "🔴"
        m.add(types.InlineKeyboardButton(f"{icon} {cat['name']}", callback_data=f"changestatus_{code}"))
    m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
    try: bot.edit_message_text("⚙️ *Category ON/OFF করুন:*\n_ক্লিক করলেই পরিবর্তন হবে_", cid, msg_id, reply_markup=m, parse_mode="Markdown")
    except: pass

def _show_payment_control(cid, msg_id):
    m = types.InlineKeyboardMarkup(row_width=1)
    b_status = "✅ ON" if SYSTEM_SETTINGS.get("pay_bkash", True) else "🔴 OFF"
    n_status = "✅ ON" if SYSTEM_SETTINGS.get("pay_nagad", True) else "🔴 OFF"
    
    m.add(types.InlineKeyboardButton(f"🟣 Bkash: {b_status}", callback_data="toggle_pay_bkash"))
    m.add(types.InlineKeyboardButton(f"🟠 Nagad: {n_status}", callback_data="toggle_pay_nagad"))
    m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_all_control"))
    
    try: bot.edit_message_text("💳 *Payment Methods Control*\n_বাটনে ক্লিক করে অফ/অন করুন_", cid, msg_id, reply_markup=m, parse_mode="Markdown")
    except: pass

def export_submissions_to_excel(cid):
    if not all_submissions:
        bot.send_message(cid, "📭 এক্সপোর্ট করার মতো কোনো ডেটা নেই।")
        return
    bot.send_message(cid, "⏳ এক্সপোর্ট ফাইল তৈরি হচ্ছে, অপেক্ষা করুন...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions Data"
        headers = ["Sub ID", "User Name", "Category", "Rate", "Submitted Qty", "Live/Approved Qty", "Total Amount (Tk)", "Account Number", "Note", "Date", "Status"]
        ws.append(headers)
        
        for sid, s in all_submissions.items():
            status = "Paid" if s.get("paid") else "Pending"
            ws.append([
                sid, s.get("user"), s.get("type_name"), s.get("rate"), s.get("qty"), 
                s.get("live_qty", 0), s.get("total", 0), s.get("account"), 
                s.get("note"), s.get("date"), status
            ])
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        filename = f"Export_Data_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        bot.send_document(cid, document=(filename, file_stream), caption="📊 *All Submissions Exported Data*", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(cid, f"❌ এক্সপোর্ট করতে সমস্যা হয়েছে: {e}")

def step_review(message, sub_id, ch_msg_id):
    cid = message.chat.id
    if not (message.text or "").isdigit():
        msg = bot.send_message(cid, "❌ শুধু সংখ্যা দিন।")
        return bot.register_next_step_handler(msg, step_review, sub_id, ch_msg_id)

    live = int(message.text)
    sub  = all_submissions.get(sub_id)
    if not sub: return

    user_data[cid] = {
        "step": "review_screenshot",
        "review_sub_id": sub_id,
        "live_qty": live,
        "ch_msg_id": ch_msg_id
    }

    amount = live * sub["rate"]
    msg = bot.send_message(cid, f"✅ Live: *{live}* টি।\n💰 Amount: *৳{amount:.2f}*\n\n📸 *এবার পেমেন্টের স্ক্রিনশট পাঠান:*\n_(স্ক্রিনশট না দিতে চাইলে `skip` বা `না` লিখুন)_", parse_mode="Markdown")
    bot.register_next_step_handler(msg, step_review_screenshot)

def step_review_screenshot(message):
    cid = message.chat.id
    if cid not in user_data or "review_sub_id" not in user_data[cid]:
        return handle_text(message)

    d = user_data[cid]
    sub_id    = d["review_sub_id"]
    live      = d["live_qty"]
    ch_msg_id = d["ch_msg_id"]
    
    sub = all_submissions.get(sub_id)
    if not sub:
        del user_data[cid]
        return bot.send_message(cid, "❌ Submission পাওয়া যায়নি।")

    fail   = max(0, sub["qty"] - live)
    amount = live * sub["rate"]
    sub.update({"live_qty": live, "total": amount, "paid": True})

    result = (
        "📊 *আপনার Result এসেছে!*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 Sub ID  : `{sub_id}`\n📁 Type    : {sub['type_name']}\n📅 Date    : {sub['date']}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Total   : {sub['qty']} pcs\n✅ Success : *{live} pcs*\n❌ Failed  : *{fail} pcs*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💰 Amount  : *{amount:.2f} Tk*\n💳 Account : `{sub['account']}`\n\n"
        "✅ *Payment প্রক্রিয়া সম্পন্ন হয়েছে!* 🎉\n"
    )

    try: 
        if message.photo:
            bot.send_photo(sub["chat_id"], message.photo[-1].file_id, caption=result, parse_mode="Markdown")
        else:
            bot.send_message(sub["chat_id"], result, parse_mode="Markdown")
    except Exception as e: 
        log.warning(f"Result send failed: {e}")

    if ch_msg_id:
        done_btn = types.InlineKeyboardMarkup()
        done_btn.add(types.InlineKeyboardButton(f"✅ Completed — {live} Paid", callback_data="already_done"))
        try: bot.edit_message_reply_markup(CHANNEL_ID, ch_msg_id, reply_markup=done_btn)
        except: pass

    bot.send_message(cid, f"✅ *#{sub_id} Result পাঠানো হয়েছে!*\nAmount: ৳{amount:.2f}", parse_mode="Markdown")
    del user_data[cid]

def step_set_notice(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    text = (message.text or "").strip()
    if text.lower() == "clear":
        SYSTEM_SETTINGS["notice_board"] = ""
        bot.send_message(cid, "✅ *Notice Board মুছে দেওয়া হয়েছে!*", parse_mode="Markdown")
    else:
        SYSTEM_SETTINGS["notice_board"] = text
        bot.send_message(cid, "✅ *Notice Board আপডেট করা হয়েছে!*\nইউজাররা এখন এটি দেখতে পাবে।", parse_mode="Markdown")
    show_admin_panel(cid)

def step_admin_search_sub(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    sub_id = (message.text or "").strip().upper()
    sub = all_submissions.get(sub_id)
    if not sub:
        bot.send_message(cid, f"❌ *'{sub_id}'* পাওয়া যায়নি।", parse_mode="Markdown")
        return
    
    status = "✅ Paid" if sub.get("paid") else "⏳ Pending"
    text = (
        f"🔎 *Submission Details: {sub_id}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 User: {sub.get('user')} | `{sub.get('chat_id')}`\n"
        f"📁 Type: {sub.get('type_name')}\n"
        f"📅 Date: {sub.get('date')}\n"
        f"📦 Submit: {sub.get('qty')} pcs\n"
        f"✅ Live: {sub.get('live_qty', 0)} pcs\n"
        f"💰 Amount: ৳{sub.get('total', 0):.2f}\n"
        f"💳 Account: `{sub.get('account')}`\n"
        f"📝 Note: {sub.get('note')}\n"
        f"📌 Status: {status}"
    )
    
    m = types.InlineKeyboardMarkup()
    if not sub.get("paid"):
        m.add(types.InlineKeyboardButton("🔍 Review এখনই করুন", callback_data=f"review_{sub_id}"))
    bot.send_message(cid, text, parse_mode="Markdown", reply_markup=m)

def step_admin_view_user_history(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    
    target_input = (message.text or "").strip()
    target_id = None
    
    if target_input.lstrip("-").isdigit():
        target_id = int(target_input)
    elif target_input.lower() in username_to_id:
        target_id = username_to_id[target_input.lower()]
    
    if not target_id:
        bot.send_message(cid, "❌ ইউজার পাওয়া যায়নি।")
        return
    
    subs = user_submissions.get(target_id, [])
    if not subs:
        bot.send_message(cid, f"📭 এই ইউজারের (`{target_id}`) কোনো সাবমিশন নেই।", parse_mode="Markdown")
        return
    
    total_earned = 0.0
    lines = [f"📋 *User History: `{target_id}`*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    for sid in reversed(subs[-20:]):
        s = all_submissions.get(sid, {})
        icon = "✅" if s.get("paid") else "⏳"
        amt = s.get("total", 0)
        total_earned += amt
        lines.append(f"{icon} `{sid}` | {s.get('type_name','?')} | ৳{amt:.2f}")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━")
    lines.append(f"💰 মোট আয়: *৳{total_earned:.2f}*")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

def step_edit_welcome(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    SYSTEM_SETTINGS["welcome_msg"] = message.text.strip()
    bot.send_message(cid, "✅ *ওয়েলকাম মেসেজ সফলভাবে আপডেট করা হয়েছে!*", parse_mode="Markdown")
    show_admin_panel(cid)

def step_ban_unban_user(message):
    cid = message.chat.id
    target_id = message.text.strip()
    if not target_id.isdigit():
        bot.send_message(cid, "❌ শুধু ইউজারের Telegram ID (সংখ্যায়) দিন।")
        return
    
    target_id = int(target_id)
    if target_id in ADMIN_IDS:
        bot.send_message(cid, "❌ অ্যাডমিনকে ব্যান করা যাবে না!")
        return

    if target_id in BANNED_USERS:
        BANNED_USERS.remove(target_id)
        bot.send_message(cid, f"✅ ইউজার `{target_id}` কে *আনব্যান* করা হয়েছে!", parse_mode="Markdown")
    else:
        BANNED_USERS.add(target_id)
        bot.send_message(cid, f"🚫 ইউজার `{target_id}` কে *ব্যান* করা হয়েছে!", parse_mode="Markdown")

def handle_db_backup(cid):
    bot.send_message(cid, "⏳ ডাটাবেস ব্যাকআপ তৈরি হচ্ছে...")
    try:
        db_dump = {
            "submissions": all_submissions,
            "user_submissions": user_submissions,
            "registered_users": list(registered_users),
            "banned_users": list(BANNED_USERS),
            "settings": SYSTEM_SETTINGS
        }
        json_data = json.dumps(db_dump, indent=4, default=str)
        file_stream = BytesIO(json_data.encode("utf-8"))
        
        filename = f"DB_Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        bot.send_document(cid, document=(filename, file_stream), caption="💾 *সম্পূর্ণ ডাটাবেস ব্যাকআপ*\n\nআপনার বটের সব সেভড ডাটা এখানে আছে।", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(cid, f"❌ ব্যাকআপ নিতে সমস্যা হয়েছে: {e}")

def step_change_min_limit(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    try:
        new_limit = int(message.text.strip())
        SYSTEM_SETTINGS["min_id_limit"] = new_limit
        bot.send_message(cid, f"✅ *সফল!*\nএখন থেকে কেউ {new_limit} টির কম আইডি জমা দিতে পারবে না।", parse_mode="Markdown")
        show_admin_panel(cid)
    except ValueError:
        msg = bot.send_message(cid, "❌ ভুল ফরম্যাট! শুধু সংখ্যা দিন (যেমন: 10):")
        bot.register_next_step_handler(msg, step_change_min_limit)

def step_msg_user_id(message):
    uid_input = (message.text or "").strip()
    target = None
    if uid_input.lstrip("-").isdigit(): target = int(uid_input)
    elif uid_input.lower() in username_to_id: target = username_to_id[uid_input.lower()]
    else: bot.send_message(message.chat.id, "❌ *পাওয়া যায়নি!*", parse_mode="Markdown"); return
    msg = bot.send_message(message.chat.id, "✍️ মেসেজ বা ছবি পাঠান:")
    bot.register_next_step_handler(msg, step_send_to_user, target)

def step_send_to_user(message, target):
    try:
        if message.photo: bot.send_photo(target, message.photo[-1].file_id, caption=message.caption or "")
        else: bot.send_message(target, f"📩 *Admin Message:*\n\n{message.text}", parse_mode="Markdown")
        bot.send_message(message.chat.id, f"✅ পাঠানো হয়েছে → `{target}`", parse_mode="Markdown")
    except Exception as e: bot.send_message(message.chat.id, f"❌ ব্যর্থ: {e}")

def step_broadcast(message):
    if message.text in MENU_BUTTONS: return handle_text(message)
    ok = fail = 0
    bot.send_message(message.chat.id, f"⏳ ব্রডকাস্ট শুরু...")
    for uid in list(registered_users):
        try:
            if message.photo: 
                bot.send_photo(uid, message.photo[-1].file_id, caption=message.caption or "")
            elif message.sticker:
                bot.send_sticker(uid, message.sticker.file_id)
            else: 
                bot.send_message(uid, message.text, parse_mode="Markdown")
            ok += 1
        except: fail += 1
        time.sleep(0.05)
    bot.send_message(message.chat.id, f"📢 *ব্রডকাস্ট সম্পন্ন!*\n✅ সফল: {ok}\n❌ ব্যর্থ: {fail}", parse_mode="Markdown")

def step_set_rate(message, code):
    try:
        new_rate = float(message.text.strip())
        CATEGORIES[code]["rate"] = new_rate
        bot.send_message(message.chat.id, f"✅ *{CATEGORIES[code]['name']}*\nনতুন rate: *{new_rate:.2f} Tk*", parse_mode="Markdown")
        show_admin_panel(message.chat.id)
    except ValueError:
        msg = bot.send_message(message.chat.id, "❌ সংখ্যা দিন:")
        bot.register_next_step_handler(msg, step_set_rate, code)

def show_history(cid):
    subs = user_submissions.get(cid, [])
    if not subs: bot.send_message(cid, "📭 এখনো কোনো সাবমিশন নেই।"); return
    lines = ["📜 *আপনার সাবমিশন হিস্টোরি*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    for sid in reversed(subs[-15:]):
        s    = all_submissions.get(sid, {})
        icon = "✅" if s.get("paid") else "⏳"
        amt  = s.get("total", 0)
        lines.append(f"{icon} `{sid}` | {s.get('type_name','?')} | ৳{amt:.2f} | {s.get('date','?')}")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

def show_admin_panel(cid):
    total  = len(all_submissions)
    paid   = sum(1 for s in all_submissions.values() if s.get("paid"))
    tk     = sum(s.get("total", 0) for s in all_submissions.values() if s.get("paid"))
    users  = len(registered_users)
    banned = len(BANNED_USERS)
    
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(types.InlineKeyboardButton("⏳ Pending", callback_data="admin_pending_subs"),
          types.InlineKeyboardButton("📋 All Subs", callback_data="admin_all_subs"))
    m.add(types.InlineKeyboardButton("💰 Rates", callback_data="admin_change_rate_menu"),
          types.InlineKeyboardButton("⚙️ Categories", callback_data="admin_change_status_menu"))
    m.add(types.InlineKeyboardButton("👁️ User History", callback_data="admin_user_history"),
          types.InlineKeyboardButton("📩 Msg User", callback_data="admin_msg_user"))
    m.add(types.InlineKeyboardButton("🔎 Search ID", callback_data="admin_search_sub"),
          types.InlineKeyboardButton("📥 Backup", callback_data="admin_export_data"))
    m.add(types.InlineKeyboardButton("🎛️ System Settings", callback_data="admin_all_control"),
          types.InlineKeyboardButton("🔄 Refresh", callback_data="admin_refresh"))
    
    admin_text = (
        "👑 *ADVANCED ADMIN CONTROL*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 Total Users : *{users}*\n"
        f"🚫 Banned      : *{banned}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📥 Submissions : *{total}*\n"
        f"✅ Paid        : *{paid}*\n"
        f"⏳ Pending     : *{total - paid}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💵 Total Paid  : *৳{tk:.2f}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    bot.send_message(cid, admin_text, reply_markup=m, parse_mode="Markdown")

@bot.message_handler(func=lambda m: True, content_types=["photo", "video", "audio", "voice", "sticker"])
def handle_media(message):
    cid = message.chat.id
    if is_banned(cid): return
    if cid in user_data and user_data[cid].get("step") == "file":
        bot.send_message(cid, "❌ একটি *.xlsx* ফাইল পাঠান।", reply_markup=cancel_btn(), parse_mode="Markdown")
    else:
        bot.send_message(cid, "❓ শুধু নিচের বাটনগুলো ব্যবহার করুন।", reply_markup=main_menu(cid))

# ════════════════════════════════════════════════════════════════
#  RUN
# ════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    log.info(f"🚀 {BOT_NAME} starting...")
    while True:
        try:
            bot.infinity_polling(timeout=60, long_polling_timeout=30, skip_pending=True)
        except Exception as e:
            log.error(f"Polling crashed: {e}")
            time.sleep(5)
